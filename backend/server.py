from fastapi import FastAPI, APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from contextlib import asynccontextmanager
import os
import logging
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime, timezone, timedelta
import uuid
from typing import List, Optional
import asyncio
import httpx
import pandas as pd
import io
from fastapi.responses import StreamingResponse
from datetime import datetime

from models import *
from auth import get_password_hash, verify_password, create_access_token, get_current_user, check_role
from telematics import get_telematics_provider, TelematicsService
from tracking_service import TrackingService
from automation import AutomationService
import cloudinary
import cloudinary.uploader
from surepass import SurepassService
import json

# Helper function to convert datetime objects to readable format

def serialize_doc(doc):
    """Convert datetime objects in a dict to readable format for JSON serialization"""
    if isinstance(doc, dict):
        result = {}
        for k, v in doc.items():
            if k == '_id':
                continue
            if isinstance(v, datetime):
                # Format as YYYY-MM-DD without time component
                result[k] = v.strftime('%Y-%m-%d')
            elif isinstance(v, dict):
                result[k] = serialize_doc(v)
            elif isinstance(v, list):
                result[k] = [serialize_doc(item) for item in v]
            elif hasattr(v, '__dict__') or (hasattr(v, 'items') and callable(v.items)):
                try:
                    result[k] = str(v)
                except:
                    result[k] = None
            else:
                result[k] = v
        return result
    elif isinstance(doc, datetime):
        return doc.strftime('%Y-%m-%d')
    elif isinstance(doc, list):
        return [serialize_doc(item) for item in doc]
    elif hasattr(doc, '__dict__') or (hasattr(doc, 'items') and callable(doc.items)):
        try:
            return str(doc)
        except:
            return None
    return doc

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET"),
    secure=True
)


MAX_FILE_SIZE = 3 * 1024 * 1024  # 5MB

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Global services
telematics_service = TelematicsService(get_telematics_provider("Simulation"))
automation_service = None
tracking_service = None

# Background task for automation
async def automation_background_task():
    """Run automation checks periodically"""
    while True:
        try:
            await automation_service.run_all_checks()
        except Exception as e:
            logging.error(f"Automation task error: {e}")
        # Run every 6 hours
        await asyncio.sleep(6 * 60 * 60)

async def tracking_background_task():
    global tracking_service

    await asyncio.sleep(2)

    while True:
        try:
            if not tracking_service:
                print("Tracking service not initialized yet...")
                await asyncio.sleep(5)
                continue

            interval_doc = await db.tracking_settings.find_one({})
            interval = interval_doc.get("interval", 60) if interval_doc else 60

            print(f"[TRACKING] Fetching data... Interval: {interval}s")

            await tracking_service.fetch_live_data()
            # for debugging
            # await tracking_service.cleanup_old_records(days=0) 
            

            await asyncio.sleep(interval)

        except Exception as e:
            print("❌ Tracking Error:", e)
            await asyncio.sleep(60)

@asynccontextmanager
async def lifespan(app: FastAPI):
    global automation_service, tracking_service

    print("🚀 Starting application...")

    automation_service = AutomationService(db)
    tracking_service = TrackingService(db)
    await tracking_service.setup_ttl_index()
    task1 = asyncio.create_task(automation_background_task())
    task2 = asyncio.create_task(tracking_background_task())

    try:
        yield
    finally:
        print("🛑 Shutting down...")

        task1.cancel()
        task2.cancel()

        try:
            await task1
        except asyncio.CancelledError:
            pass

        try:
            await task2
        except asyncio.CancelledError:
            pass

        client.close()

app = FastAPI(lifespan=lifespan)
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

surepass_service = SurepassService()
# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    # Check if user exists
    existing = await db.users.find_one({"email": user_data.email, "is_deleted": False})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    user = User(
        id=user_id,
        email=user_data.email,
        full_name=user_data.full_name,
        role=user_data.role,
        is_active=user_data.is_active,
        hashed_password=get_password_hash(user_data.password)
    )
    
    user_dict = user.model_dump()
    user_dict["created_at"] = user_dict["created_at"].isoformat()
    user_dict["updated_at"] = user_dict["updated_at"].isoformat()
    
    await db.users.insert_one(user_dict)
    
    return {"id": user_id, "email": user.email, "role": user.role}

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email, "is_deleted": False})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(credentials.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account is inactive")
    
    token = create_access_token({"user_id": user["id"], "email": user["email"], "role": user["role"]})
    
    return {"access_token": token, "token_type": "bearer", "user": {"id": user["id"], "email": user["email"], "full_name": user["full_name"], "role": user["role"]}}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["user_id"], "is_deleted": False}, {"_id": 0, "hashed_password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

# ==================== PROPERTY ROUTES ====================

@api_router.post("/properties")
async def create_property(property_data: PropertyCreate, current_user: dict = Depends(get_current_user)):
    property_id = str(uuid.uuid4())
    prop = Property(id=property_id, **property_data.model_dump(), created_by=current_user["user_id"])
    
    prop_dict = prop.model_dump()
    prop_dict["created_at"] = prop_dict["created_at"].isoformat()
    prop_dict["updated_at"] = prop_dict["updated_at"].isoformat()
    
    await db.properties.insert_one(prop_dict)
    return serialize_doc(prop_dict)

@api_router.get("/properties")
async def get_properties(
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    properties = await db.properties.find({"is_deleted": False}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    count = await db.properties.count_documents({"is_deleted": False})
    return {"data": properties, "total": count, "skip": skip, "limit": limit}

@api_router.get("/properties/{property_id}")
async def get_property(property_id: str, current_user: dict = Depends(get_current_user)):
    prop = await db.properties.find_one({"id": property_id, "is_deleted": False}, {"_id": 0})
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")
    return prop

@api_router.put("/properties/{property_id}")
async def update_property(property_id: str, property_data: PropertyCreate, current_user: dict = Depends(get_current_user)):
    update_data = property_data.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["user_id"]
    
    result = await db.properties.update_one({"id": property_id, "is_deleted": False}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Property not found")
    return {"message": "Property updated"}

@api_router.delete("/properties/{property_id}")
async def delete_property(property_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.properties.update_one(
        {"id": property_id},
        {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Property not found")
    return {"message": "Property deleted"}

# ==================== PROPERTY TAX ROUTES ====================

@api_router.post("/property-taxes")
async def create_property_tax(tax_data: PropertyTaxCreate, current_user: dict = Depends(get_current_user)):
    tax_id = str(uuid.uuid4())
    tax = PropertyTax(id=tax_id, **tax_data.model_dump(), created_by=current_user["user_id"])
    
    tax_dict = tax.model_dump()
    tax_dict["created_at"] = tax_dict["created_at"].isoformat()
    tax_dict["updated_at"] = tax_dict["updated_at"].isoformat()
    tax_dict["issue_date"] = tax_dict["issue_date"].isoformat()
    tax_dict["expiry_date"] = tax_dict["expiry_date"].isoformat()
    if tax_dict.get("payment_date"):
        tax_dict["payment_date"] = tax_dict["payment_date"].isoformat()
    
    await db.property_taxes.insert_one(tax_dict)
    return serialize_doc(tax_dict)

@api_router.get("/property-taxes")
async def get_property_taxes(
    property_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": False}
    if property_id:
        query["property_id"] = property_id
    
    taxes = await db.property_taxes.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    count = await db.property_taxes.count_documents(query)
    return {"data": taxes, "total": count}

@api_router.get("/property-taxes/{tax_id}")
async def get_property_tax(tax_id: str, current_user: dict = Depends(get_current_user)):
    tax = await db.property_taxes.find_one({"id": tax_id, "is_deleted": False}, {"_id": 0})
    if not tax:
        raise HTTPException(status_code=404, detail="Property tax not found")
    return tax

@api_router.put("/property-taxes/{tax_id}")
async def update_property_tax(tax_id: str, tax_data: PropertyTaxCreate, current_user: dict = Depends(get_current_user)):
    update_data = tax_data.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["user_id"]
    update_data["issue_date"] = update_data["issue_date"].isoformat()
    update_data["expiry_date"] = update_data["expiry_date"].isoformat()
    if update_data.get("payment_date"):
        update_data["payment_date"] = update_data["payment_date"].isoformat()
    
    result = await db.property_taxes.update_one({"id": tax_id, "is_deleted": False}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Property tax not found")
    return {"message": "Property tax updated"}

@api_router.delete("/property-taxes/{tax_id}")
async def delete_property_tax(tax_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.property_taxes.update_one(
        {"id": tax_id},
        {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Property tax not found")
    return {"message": "Property tax deleted"}

# ==================== ELECTRICITY BILL ROUTES ====================

@api_router.post("/electricity-bills/fetch-from-surepass")
async def fetch_electricity_bill_from_surepass(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Fetch electricity bill details from Surepass API using consumer ID
    """
    consumer_id = data.get("consumer_id", "").strip()
    operator_code = data.get("operator_code", "").upper().strip()
    property_id = data.get("property_id")
    
    if not consumer_id:
        raise HTTPException(status_code=400, detail="Consumer ID is required")
    
    if not operator_code:
        raise HTTPException(status_code=400, detail="Operator code is required (e.g., 'MH' for Maharashtra)")
    
    if not property_id:
        raise HTTPException(status_code=400, detail="Property ID is required")
    
    # FIX: Remove the 3-character validation - operator codes can be 2-4 characters
    # Instead, just validate it's not empty
    if len(operator_code) < 2:
        raise HTTPException(status_code=400, detail="Operator code must be at least 2 characters")
    
    # Verify property exists
    property = await db.properties.find_one({"id": property_id, "is_deleted": False})
    if not property:
        raise HTTPException(status_code=404, detail="Property not found")
    
    logger.info(f"Fetching electricity bill for consumer: {consumer_id}, operator: {operator_code}")
    
    # Call Surepass API
    result = await surepass_service.fetch_electricity_bill(consumer_id, operator_code)
    
    if not result["success"]:
        error_detail = result.get("error", "Failed to fetch electricity bill details")
        logger.error(f"Surepass API error: {error_detail}")
        raise HTTPException(status_code=400, detail=error_detail)
    
    bill_data = result["data"]
    
    # Log the API call
    try:
        log_entry = {
            "id": str(uuid.uuid4()),
            "property_id": property_id,
            "consumer_id": consumer_id,
            "operator_code": operator_code,
            "request_timestamp": datetime.now(timezone.utc).isoformat(),
            "response_data": bill_data,
            "is_successful": True,
            "created_by": current_user["user_id"]
        }
        await db.electricity_verification_logs.insert_one(log_entry)
    except Exception as e:
        logger.warning(f"Could not log electricity verification: {e}")
    
    return {
        "success": True,
        "property_id": property_id,
        "property_name": property["name"],
        "bill_data": bill_data,
        "message": "Electricity bill details fetched successfully"
    }

@api_router.post("/electricity-bills/save-from-surepass")
async def save_electricity_bill_from_surepass(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Save electricity bill from Surepass data to the database
    """
    property_id = data.get("property_id")
    bill_data = data.get("bill_data", {})
    
    if not property_id:
        raise HTTPException(status_code=400, detail="Property ID is required")
    
    if not bill_data:
        raise HTTPException(status_code=400, detail="Bill data is required")
    
    # Verify property exists
    property = await db.properties.find_one({"id": property_id, "is_deleted": False})
    if not property:
        raise HTTPException(status_code=404, detail="Property not found")
    
    # Get billing period (default to current month if not provided)
    billing_period_start = data.get("billing_period_start")
    billing_period_end = data.get("billing_period_end")
    due_date = data.get("due_date")
    
    # Get payment status and payment date
    status = data.get("status", "Unpaid")
    payment_date = data.get("payment_date")
    
    # Parse dates or use defaults
    today = datetime.now(timezone.utc)
    if not billing_period_start:
        billing_period_start = today.replace(day=1).isoformat()
    if not billing_period_end:
        # Last day of current month
        if today.month == 12:
            next_month = today.replace(year=today.year + 1, month=1, day=1)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
        billing_period_end = (next_month - timedelta(days=1)).isoformat()
    if not due_date:
        due_date = (today + timedelta(days=15)).isoformat()
    
    # Validate payment date if status is Paid
    if status == "Paid" and not payment_date:
        raise HTTPException(status_code=400, detail="Payment date is required when status is Paid")
    
    # Create electricity bill
    bill_id = str(uuid.uuid4())
    
    # Parse the date strings
    billing_period_start_dt = datetime.fromisoformat(billing_period_start) if isinstance(billing_period_start, str) else billing_period_start
    billing_period_end_dt = datetime.fromisoformat(billing_period_end) if isinstance(billing_period_end, str) else billing_period_end
    due_date_dt = datetime.fromisoformat(due_date) if isinstance(due_date, str) else due_date
    
    bill_dict = {
        "id": bill_id,
        "property_id": property_id,
        "billing_period_start": billing_period_start_dt.isoformat(),
        "billing_period_end": billing_period_end_dt.isoformat(),
        "previous_reading": data.get("previous_reading", 0),
        "current_reading": data.get("current_reading", 0),
        "units_consumed": data.get("units_consumed", 0),
        "slab_charges": data.get("slab_charges", 0),
        "fixed_charges": data.get("fixed_charges", 0),
        "taxes": data.get("taxes", 0),
        "penalty": data.get("penalty", 0),
        "total_amount": bill_data.get("bill_amount", 0),
        "due_date": due_date_dt.isoformat(),
        "payment_date": payment_date if payment_date else None,
        "status": status,
        "bill_url": bill_data.get("document_link"),
        "phone_number": bill_data.get("mobile"),
        "consumer_id": bill_data.get("consumer_id"),
        "operator_code": bill_data.get("operator_code"),
        "consumer_name": bill_data.get("full_name"),
        "consumer_address": bill_data.get("address"),
        "source": "surepass",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["user_id"],
        "is_deleted": False
    }
    
    # If payment date is provided, parse it
    if payment_date:
        payment_date_dt = datetime.fromisoformat(payment_date) if isinstance(payment_date, str) else payment_date
        bill_dict["payment_date"] = payment_date_dt.isoformat()
    
    await db.electricity_bills.insert_one(bill_dict)
    
    return {
        "success": True,
        "bill_id": bill_id,
        "message": "Electricity bill saved successfully"
    }

@api_router.get("/electricity-bills/operator-codes")
async def get_electricity_operator_codes(
    current_user: dict = Depends(get_current_user)
):
    """
    Get list of electricity operator codes from Surepass
    """
    operator_codes = [
        {"code": "AS", "name": "Assam", "state": "Assam"},
        {"code": "AJ", "name": "Rajasthan (Ajmer)", "state": "Rajasthan"},
        {"code": "APS", "name": "Andhra Pradesh", "state": "Andhra Pradesh"},
        {"code": "BG", "name": "West Bengal", "state": "West Bengal"},
        {"code": "BS", "name": "Karnataka (BESCOM)", "state": "Karnataka"},
        {"code": "BR", "name": "BSES Rajdhani Delhi", "state": "Delhi"},
        {"code": "BY", "name": "BSES Yamuna Delhi", "state": "Delhi"},
        {"code": "CD", "name": "Chandigarh", "state": "Chandigarh"},
        {"code": "CH", "name": "Chhattisgarh", "state": "Chhattisgarh"},
        {"code": "CS", "name": "Karnataka (CESCOM)", "state": "Karnataka"},
        {"code": "DDM", "name": "Daman and Diu", "state": "Daman and Diu"},
        {"code": "DG", "name": "Gujarat (DGVCL)", "state": "Gujarat"},
        {"code": "DH", "name": "Haryana (DHBVN)", "state": "Haryana"},
        {"code": "DL", "name": "Delhi (Tata Power)", "state": "Delhi"},
        {"code": "DNH", "name": "Dadra and Nagar Haveli", "state": "Dadra and Nagar Haveli"},
        {"code": "GO", "name": "Goa", "state": "Goa"},
        {"code": "GS", "name": "Karnataka (GESCOM)", "state": "Karnataka"},
        {"code": "HP", "name": "Himachal Pradesh", "state": "Himachal Pradesh"},
        {"code": "HS", "name": "Karnataka (HESCOM)", "state": "Karnataka"},
        {"code": "JD", "name": "Rajasthan (Jodhpur)", "state": "Rajasthan"},
        {"code": "JP", "name": "Rajasthan (Jaipur)", "state": "Rajasthan"},
        {"code": "KP", "name": "Kanpur", "state": "Uttar Pradesh"},
        {"code": "KC", "name": "West Bengal (Kolkata)", "state": "West Bengal"},
        {"code": "MG", "name": "Gujarat (MGVCL)", "state": "Gujarat"},
        {"code": "MH", "name": "Maharashtra (MSEDCL)", "state": "Maharashtra"},
        {"code": "MPC", "name": "Madhya Pradesh (Central)", "state": "Madhya Pradesh"},
        {"code": "MPE", "name": "Madhya Pradesh (East)", "state": "Madhya Pradesh"},
        {"code": "MPW", "name": "Madhya Pradesh (West)", "state": "Madhya Pradesh"},
        {"code": "MS", "name": "Karnataka (MESCOM)", "state": "Karnataka"},
        {"code": "MU", "name": "Mumbai (Tata Power)", "state": "Maharashtra"},
        {"code": "AEML", "name": "Mumbai (Adani Electricity)", "state": "Maharashtra"},
        {"code": "NB", "name": "Bihar (North Bihar)", "state": "Bihar"},
        {"code": "NG", "name": "Nagaland", "state": "Nagaland"},
        {"code": "OD", "name": "Odisha", "state": "Odisha"},
        {"code": "PG", "name": "Gujarat (PGVCL)", "state": "Gujarat"},
        {"code": "PN", "name": "Punjab", "state": "Punjab"},
        {"code": "SB", "name": "Bihar (South Bihar)", "state": "Bihar"},
        {"code": "TG", "name": "Torrent Power (Gujarat)", "state": "Gujarat"},
        {"code": "TL", "name": "Telangana", "state": "Telangana"},
        {"code": "TN", "name": "Tamil Nadu", "state": "Tamil Nadu"},
        {"code": "TR", "name": "Tripura", "state": "Tripura"},
        {"code": "UG", "name": "Gujarat (UGVCL)", "state": "Gujarat"},
        {"code": "UH", "name": "Haryana (UHBVN)", "state": "Haryana"},
        {"code": "UK", "name": "Uttarakhand", "state": "Uttarakhand"},
        {"code": "UP", "name": "Uttar Pradesh (UPPCL)", "state": "Uttar Pradesh"},
        {"code": "PUP", "name": "Purvanchal Uttar Pradesh", "state": "Uttar Pradesh"},
        {"code": "MUP", "name": "Madhyanchal Uttar Pradesh", "state": "Uttar Pradesh"},
    ]
    
    # Sort by state name for better UX
    operator_codes.sort(key=lambda x: x["state"])
    
    return {"data": operator_codes}

@api_router.post("/electricity-bills")
async def create_electricity_bill(bill_data: ElectricityBillCreate, current_user: dict = Depends(get_current_user)):
    bill_id = str(uuid.uuid4())
    bill = ElectricityBill(id=bill_id, **bill_data.model_dump(), created_by=current_user["user_id"])
    
    bill_dict = bill.model_dump()
    bill_dict["created_at"] = bill_dict["created_at"].isoformat()
    bill_dict["updated_at"] = bill_dict["updated_at"].isoformat()
    bill_dict["billing_period_start"] = bill_dict["billing_period_start"].isoformat()
    bill_dict["billing_period_end"] = bill_dict["billing_period_end"].isoformat()
    bill_dict["due_date"] = bill_dict["due_date"].isoformat()
    if bill_dict.get("payment_date"):
        bill_dict["payment_date"] = bill_dict["payment_date"].isoformat()
    
    await db.electricity_bills.insert_one(bill_dict)
    return serialize_doc(bill_dict)

@api_router.get("/electricity-bills")
async def get_electricity_bills(
    property_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": False}
    if property_id:
        query["property_id"] = property_id
    
    bills = await db.electricity_bills.find(query, {"_id": 0}).sort("billing_period_start", -1).skip(skip).limit(limit).to_list(limit)
    count = await db.electricity_bills.count_documents(query)
    return {"data": bills, "total": count}

@api_router.get("/electricity-bills/{bill_id}")
async def get_electricity_bill(bill_id: str, current_user: dict = Depends(get_current_user)):
    bill = await db.electricity_bills.find_one({"id": bill_id, "is_deleted": False}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Electricity bill not found")
    return bill

# Add this after the GET /electricity-bills/{bill_id} endpoint and before the DELETE endpoint

@api_router.put("/electricity-bills/{bill_id}")
async def update_electricity_bill(bill_id: str, bill_data: ElectricityBillCreate, current_user: dict = Depends(get_current_user)):
    update_data = bill_data.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["user_id"]
    update_data["billing_period_start"] = update_data["billing_period_start"].isoformat()
    update_data["billing_period_end"] = update_data["billing_period_end"].isoformat()
    update_data["due_date"] = update_data["due_date"].isoformat()
    if update_data.get("payment_date"):
        update_data["payment_date"] = update_data["payment_date"].isoformat()
    
    result = await db.electricity_bills.update_one(
        {"id": bill_id, "is_deleted": False}, 
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Electricity bill not found")
    
    return {"message": "Electricity bill updated successfully"}    

@api_router.delete("/electricity-bills/{bill_id}")
async def delete_electricity_bill(bill_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.electricity_bills.update_one(
        {"id": bill_id},
        {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Electricity bill not found")
    return {"message": "Electricity bill deleted"}

# ==================== SOLAR METER ROUTES ====================

@api_router.post("/solar-meters")
async def create_solar_meter(meter_data: SolarMeterCreate, current_user: dict = Depends(get_current_user)):
    meter_id = str(uuid.uuid4())
    meter = SolarMeter(id=meter_id, **meter_data.model_dump(), created_by=current_user["user_id"])
    
    # Check reconciliation: Solar Generated - Exported + Imported = Billable Units
    expected_billable = meter.units_generated - meter.exported_to_grid + meter.imported_from_grid
    if abs(expected_billable - meter.billable_units) > 1:  # 1 unit tolerance
        meter.reconciliation_flag = True
    
    meter_dict = meter.model_dump()
    meter_dict["created_at"] = meter_dict["created_at"].isoformat()
    meter_dict["updated_at"] = meter_dict["updated_at"].isoformat()
    meter_dict["billing_period_start"] = meter_dict["billing_period_start"].isoformat()
    meter_dict["billing_period_end"] = meter_dict["billing_period_end"].isoformat()
    
    await db.solar_meters.insert_one(meter_dict)
    return serialize_doc(meter_dict)

@api_router.get("/solar-meters")
async def get_solar_meters(
    property_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": False}
    if property_id:
        query["property_id"] = property_id
    
    meters = await db.solar_meters.find(query, {"_id": 0}).sort("billing_period_start", -1).skip(skip).limit(limit).to_list(limit)
    count = await db.solar_meters.count_documents(query)
    return {"data": meters, "total": count}

@api_router.get("/solar-meters/{meter_id}")
async def get_solar_meter(meter_id: str, current_user: dict = Depends(get_current_user)):
    meter = await db.solar_meters.find_one({"id": meter_id, "is_deleted": False}, {"_id": 0})
    if not meter:
        raise HTTPException(status_code=404, detail="Solar meter record not found")
    return meter

# Add this after the GET /solar-meters/{meter_id} endpoint and before the DELETE endpoint

@api_router.put("/solar-meters/{meter_id}")
async def update_solar_meter(meter_id: str, meter_data: SolarMeterCreate, current_user: dict = Depends(get_current_user)):
    update_data = meter_data.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["user_id"]
    update_data["billing_period_start"] = update_data["billing_period_start"].isoformat()
    update_data["billing_period_end"] = update_data["billing_period_end"].isoformat()
    
    # Recalculate reconciliation flag
    expected_billable = update_data["units_generated"] - update_data["exported_to_grid"] + update_data["imported_from_grid"]
    update_data["reconciliation_flag"] = abs(expected_billable - update_data["billable_units"]) > 1
    
    result = await db.solar_meters.update_one(
        {"id": meter_id, "is_deleted": False}, 
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Solar meter record not found")
    
    return {"message": "Solar meter record updated successfully"}    

@api_router.delete("/solar-meters/{meter_id}")
async def delete_solar_meter(meter_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.solar_meters.update_one(
        {"id": meter_id},
        {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Solar meter record not found")
    return {"message": "Solar meter record deleted"}

# ==================== GAS BILL ROUTES ====================

@api_router.post("/gas-bills")
async def create_gas_bill(bill_data: GasBillCreate, current_user: dict = Depends(get_current_user)):
    bill_id = str(uuid.uuid4())
    bill = GasBill(id=bill_id, **bill_data.model_dump(), created_by=current_user["user_id"])
    
    bill_dict = bill.model_dump()
    bill_dict["created_at"] = bill_dict["created_at"].isoformat()
    bill_dict["updated_at"] = bill_dict["updated_at"].isoformat()
    bill_dict["billing_period_start"] = bill_dict["billing_period_start"].isoformat()
    bill_dict["billing_period_end"] = bill_dict["billing_period_end"].isoformat()
    bill_dict["due_date"] = bill_dict["due_date"].isoformat()
    if bill_dict.get("payment_date"):
        bill_dict["payment_date"] = bill_dict["payment_date"].isoformat()
    
    await db.gas_bills.insert_one(bill_dict)
    return serialize_doc(bill_dict)

@api_router.get("/gas-bills")
async def get_gas_bills(
    property_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": False}
    if property_id:
        query["property_id"] = property_id
    
    bills = await db.gas_bills.find(query, {"_id": 0}).sort("billing_period_start", -1).skip(skip).limit(limit).to_list(limit)
    count = await db.gas_bills.count_documents(query)
    return {"data": bills, "total": count}

@api_router.get("/gas-bills/{bill_id}")
async def get_gas_bill(bill_id: str, current_user: dict = Depends(get_current_user)):
    bill = await db.gas_bills.find_one({"id": bill_id, "is_deleted": False}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Gas bill not found")
    return bill

@api_router.put("/gas-bills/{bill_id}")
async def update_gas_bill(bill_id: str, bill_data: GasBillCreate, current_user: dict = Depends(get_current_user)):
    """
    Update an existing gas bill
    """
    update_data = bill_data.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["user_id"]
    update_data["billing_period_start"] = update_data["billing_period_start"].isoformat()
    update_data["billing_period_end"] = update_data["billing_period_end"].isoformat()
    update_data["due_date"] = update_data["due_date"].isoformat()
    if update_data.get("payment_date"):
        update_data["payment_date"] = update_data["payment_date"].isoformat()
    
    result = await db.gas_bills.update_one(
        {"id": bill_id, "is_deleted": False}, 
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Gas bill not found")
    
    return {"message": "Gas bill updated successfully"}


@api_router.delete("/gas-bills/{bill_id}")
async def delete_gas_bill(bill_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.gas_bills.update_one(
        {"id": bill_id},
        {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Gas bill not found")
    return {"message": "Gas bill deleted"}

@api_router.post("/gas-bills/fetch-from-surepass")
async def fetch_gas_bill_from_surepass(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Fetch gas connection details from Surepass API using mobile number and provider
    """
    mobile_number = data.get("mobile_number", "").strip()
    provider_name = data.get("provider_name", "").lower().strip()
    property_id = data.get("property_id")
    
    if not mobile_number:
        raise HTTPException(status_code=400, detail="Mobile number is required")
    
    if len(mobile_number) != 10:
        raise HTTPException(status_code=400, detail="Mobile number must be 10 digits")
    
    if not provider_name:
        raise HTTPException(status_code=400, detail="Provider name is required (e.g., 'indane', 'bharat_gas', 'hp_gas')")
    
    if not property_id:
        raise HTTPException(status_code=400, detail="Property ID is required")
    
    # Verify property exists
    property = await db.properties.find_one({"id": property_id, "is_deleted": False})
    if not property:
        raise HTTPException(status_code=404, detail="Property not found")
    
    logger.info(f"Fetching gas connection for mobile: {mobile_number}, provider: {provider_name}")
    
    # Call Surepass API
    result = await surepass_service.fetch_gas_bill(mobile_number, provider_name)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result.get("error", "Failed to fetch gas connection details"))
    
    gas_data = result["data"]
    
    # Log the API call
    try:
        log_entry = {
            "id": str(uuid.uuid4()),
            "property_id": property_id,
            "mobile_number": mobile_number,
            "provider_name": provider_name,
            "request_timestamp": datetime.now(timezone.utc).isoformat(),
            "response_data": gas_data,
            "is_successful": True,
            "created_by": current_user["user_id"]
        }
        await db.gas_verification_logs.insert_one(log_entry)
    except Exception as e:
        logger.warning(f"Could not log gas verification: {e}")
    
    return {
        "success": True,
        "property_id": property_id,
        "property_name": property["name"],
        "gas_data": gas_data,
        "message": "Gas connection details fetched successfully"
    }

@api_router.post("/gas-bills/save-from-surepass")
async def save_gas_bill_from_surepass(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Save gas bill from Surepass data to the database
    """
    property_id = data.get("property_id")
    gas_data = data.get("gas_data", {})
    
    if not property_id:
        raise HTTPException(status_code=400, detail="Property ID is required")
    
    if not gas_data:
        raise HTTPException(status_code=400, detail="Gas data is required")
    
    # Verify property exists
    property = await db.properties.find_one({"id": property_id, "is_deleted": False})
    if not property:
        raise HTTPException(status_code=404, detail="Property not found")
    
    # Get billing period (default to current month if not provided)
    billing_period_start = data.get("billing_period_start")
    billing_period_end = data.get("billing_period_end")
    due_date = data.get("due_date")
    
    # Get payment status and payment date
    status = data.get("status", "Unpaid")
    payment_date = data.get("payment_date")
    
    # Parse dates or use defaults
    today = datetime.now(timezone.utc)
    if not billing_period_start:
        billing_period_start = today.replace(day=1).isoformat()
    if not billing_period_end:
        # Last day of current month
        if today.month == 12:
            next_month = today.replace(year=today.year + 1, month=1, day=1)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
        billing_period_end = (next_month - timedelta(days=1)).isoformat()
    if not due_date:
        due_date = (today + timedelta(days=15)).isoformat()
    
    # Validate payment date if status is Paid
    if status == "Paid" and not payment_date:
        raise HTTPException(status_code=400, detail="Payment date is required when status is Paid")
    
    # Create gas bill
    bill_id = str(uuid.uuid4())
    
    # Parse the date strings
    billing_period_start_dt = datetime.fromisoformat(billing_period_start) if isinstance(billing_period_start, str) else billing_period_start
    billing_period_end_dt = datetime.fromisoformat(billing_period_end) if isinstance(billing_period_end, str) else billing_period_end
    due_date_dt = datetime.fromisoformat(due_date) if isinstance(due_date, str) else due_date
    
    bill_dict = {
        "id": bill_id,
        "property_id": property_id,
        "billing_period_start": billing_period_start_dt.isoformat(),
        "billing_period_end": billing_period_end_dt.isoformat(),
        "units_consumed": data.get("units_consumed", 0),
        "rate_per_unit": data.get("rate_per_unit", 0),
        "fixed_charges": data.get("fixed_charges", 0),
        "total_bill": data.get("total_bill", 0),
        "due_date": due_date_dt.isoformat(),
        "status": status,
        "vendor": gas_data.get("provider_name", "Unknown"),
        "bill_url": None,
        "phone_number": gas_data.get("mobile_number"),
        "consumer_id": gas_data.get("consumer_id"),
        "consumer_number": gas_data.get("consumer_number"),
        "consumer_name": gas_data.get("consumer_name"),
        "consumer_address": gas_data.get("address"),
        "distributor_name": gas_data.get("distributor_name"),
        "distributor_code": gas_data.get("distributor_code"),
        "consumer_status": gas_data.get("consumer_status"),
        "consumer_type": gas_data.get("consumer_type"),
        "source": "surepass",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["user_id"],
        "is_deleted": False
    }
    
    # If payment date is provided, parse it
    if payment_date and status == "Paid":
        payment_date_dt = datetime.fromisoformat(payment_date) if isinstance(payment_date, str) else payment_date
        bill_dict["payment_date"] = payment_date_dt.isoformat()
    
    await db.gas_bills.insert_one(bill_dict)
    
    return {
        "success": True,
        "bill_id": bill_id,
        "message": "Gas bill saved successfully"
    }
@api_router.get("/gas-bills/provider-list")
async def get_gas_providers(
    current_user: dict = Depends(get_current_user)
):
    """
    Get list of gas providers
    """
    providers = [
        {"code": "indane", "name": "Indane (Indian Oil)", "type": "LPG"},
        {"code": "bharat_gas", "name": "Bharat Gas (BPCL)", "type": "LPG"},
        {"code": "hp_gas", "name": "HP Gas (Hindustan Petroleum)", "type": "LPG"},
        {"code": "adani_gas", "name": "Adani Gas", "type": "PNG"},
        {"code": "mahanagar_gas", "name": "Mahanagar Gas (MGL)", "type": "PNG"},
        {"code": "gujarat_gas", "name": "Gujarat Gas", "type": "PNG"},
    ]
    return {"data": providers}

# ==================== WATER BILL ROUTES ====================

@api_router.post("/water-bills")
async def create_water_bill(bill_data: WaterBillCreate, current_user: dict = Depends(get_current_user)):
    bill_id = str(uuid.uuid4())
    bill = WaterBill(id=bill_id, **bill_data.model_dump(), created_by=current_user["user_id"])
    
    bill_dict = bill.model_dump()
    bill_dict["created_at"] = bill_dict["created_at"].isoformat()
    bill_dict["updated_at"] = bill_dict["updated_at"].isoformat()
    bill_dict["billing_period_start"] = bill_dict["billing_period_start"].isoformat()
    bill_dict["billing_period_end"] = bill_dict["billing_period_end"].isoformat()
    bill_dict["due_date"] = bill_dict["due_date"].isoformat()
    if bill_dict.get("payment_date"):
        bill_dict["payment_date"] = bill_dict["payment_date"].isoformat()
    
    await db.water_bills.insert_one(bill_dict)
    return serialize_doc(bill_dict)

@api_router.get("/water-bills")
async def get_water_bills(
    property_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": False}
    if property_id:
        query["property_id"] = property_id
    
    bills = await db.water_bills.find(query, {"_id": 0}).sort("billing_period_start", -1).skip(skip).limit(limit).to_list(limit)
    count = await db.water_bills.count_documents(query)
    return {"data": bills, "total": count}

@api_router.get("/water-bills/{bill_id}")
async def get_water_bill(bill_id: str, current_user: dict = Depends(get_current_user)):
    bill = await db.water_bills.find_one({"id": bill_id, "is_deleted": False}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Water bill not found")
    return bill

@api_router.delete("/water-bills/{bill_id}")
async def delete_water_bill(bill_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.water_bills.update_one(
        {"id": bill_id},
        {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Water bill not found")
    return {"message": "Water bill deleted"}

# ==================== VEHICLE ROUTES ====================

@api_router.post("/vehicles")
# async def create_vehicle(vehicle_data: VehicleCreate, current_user: dict = Depends(get_current_user)):
#     vehicle_id = str(uuid.uuid4())
#     vehicle = Vehicle(id=vehicle_id, **vehicle_data.model_dump(), created_by=current_user["user_id"])
    
#     vehicle_dict = vehicle.model_dump()
#     vehicle_dict["created_at"] = vehicle_dict["created_at"].isoformat()
#     vehicle_dict["updated_at"] = vehicle_dict["updated_at"].isoformat()
#     if vehicle_dict.get("date_of_registration"):
#         vehicle_dict["date_of_registration"] = vehicle_dict["date_of_registration"].isoformat()
    
#     await db.vehicles.insert_one(vehicle_dict)
#     return serialize_doc(vehicle_dict)
async def create_vehicle(vehicle_data: VehicleCreate, current_user: dict = Depends(get_current_user)):
    vehicle_id = str(uuid.uuid4())

    if vehicle_data.imei:
        existing = await db.vehicles.find_one({
            "imei": vehicle_data.imei,
            "is_deleted": False
        })
        if existing:
            raise HTTPException(status_code=400, detail="IMEI already exists")

    vehicle = Vehicle(
        id=vehicle_id,
        **vehicle_data.model_dump(),
        created_by=current_user["user_id"]
    )
    
    vehicle_dict = vehicle.model_dump()
    vehicle_dict["created_at"] = vehicle_dict["created_at"].isoformat()
    vehicle_dict["updated_at"] = vehicle_dict["updated_at"].isoformat()

    if vehicle_dict.get("date_of_registration"):
        vehicle_dict["date_of_registration"] = vehicle_dict["date_of_registration"].isoformat()

    if vehicle_dict.get("imei") and len(vehicle_dict["imei"]) < 10:
        raise HTTPException(status_code=400, detail="Invalid IMEI")

    await db.vehicles.insert_one(vehicle_dict)

    return serialize_doc(vehicle_dict)
@api_router.get("/vehicles")
async def get_vehicles(
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    vehicles = await db.vehicles.find({"is_deleted": False}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    count = await db.vehicles.count_documents({"is_deleted": False})
    return {"data": vehicles, "total": count}

@api_router.get("/vehicles/{vehicle_id}")
async def get_vehicle(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    return vehicle

@api_router.put("/vehicles/{vehicle_id}")
async def update_vehicle(
    vehicle_id: str,
    vehicle_data: VehicleCreate,
    current_user: dict = Depends(get_current_user)
):
    update_data = vehicle_data.model_dump()

    # 🔴 1. Check IMEI uniqueness (if updating IMEI)
    if update_data.get("imei"):
        existing = await db.vehicles.find_one({
            "imei": update_data["imei"],
            "id": {"$ne": vehicle_id},
            "is_deleted": False
        })
        if existing:
            raise HTTPException(status_code=400, detail="IMEI already exists")

        # 🔴 Optional: basic validation
        if len(update_data["imei"]) < 10:
            raise HTTPException(status_code=400, detail="Invalid IMEI")

    # 🔴 2. Prevent accidental null overwrite
    update_data = {k: v for k, v in update_data.items() if v is not None}

    # 🔴 3. Add metadata
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["user_id"]

    # 🔴 4. Date conversion
    if update_data.get("date_of_registration"):
        update_data["date_of_registration"] = update_data["date_of_registration"].isoformat()

    result = await db.vehicles.update_one(
        {"id": vehicle_id, "is_deleted": False},
        {"$set": update_data}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    return {"message": "Vehicle updated"}
@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    return {"message": "Vehicle deleted"}

@api_router.get("/vehicles/{vehicle_id}/full-report")
async def get_vehicle_full_report(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    """Get comprehensive vehicle report with all related data"""
    vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Get all related data
    documents = await db.vehicle_documents.find(
        {
            "vehicle_id": vehicle_id,
            "$or": [
                {"is_deleted": {"$exists": False}},
                {"is_deleted": False}
            ]
        }, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    challans = await db.challans.find(
        {"vehicle_id": vehicle_id, "is_deleted": False}, {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    services = await db.service_records.find(
        {"vehicle_id": vehicle_id, "is_deleted": False}, {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    accidents = await db.accidents.find(
        {"vehicle_id": vehicle_id, "is_deleted": False}, {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Calculate summary stats
    total_challan_amount = sum(c.get("amount", 0) for c in challans)
    unpaid_challans = len([c for c in challans if c.get("status") == "Unpaid"])
    total_service_cost = sum(s.get("total_cost", 0) for s in services)
    active_documents = len([d for d in documents if d.get("status") == "Active"])
    expired_documents = len([d for d in documents if d.get("status") == "Expired"])
    
    return {
        "vehicle": vehicle,
        "documents": documents,
        "challans": challans,
        "services": services,
        "accidents": accidents,
        "summary": {
            "total_documents": len(documents),
            "active_documents": active_documents,
            "expired_documents": expired_documents,
            "total_challans": len(challans),
            "unpaid_challans": unpaid_challans,
            "total_challan_amount": total_challan_amount,
            "total_services": len(services),
            "total_service_cost": total_service_cost,
            "total_accidents": len(accidents)
        }
    }

@api_router.get("/vehicles/export/csv")
async def export_vehicles_csv(current_user: dict = Depends(get_current_user)):
    """Export all vehicles as CSV with comprehensive fields"""
    vehicles = await db.vehicles.find({"is_deleted": False}, {"_id": 0}).to_list(1000)
    
    # Comprehensive CSV headers with user-friendly names
    headers = [
        "Registration Number", "Vehicle Type", "Brand", "Model", "Year", 
        "Chassis Number", "Engine Number", "Color", "Fuel Type", 
        "Average Mileage (km/l)", "Tank Capacity (L)", "Seating Capacity", 
        "Owner Name", "File Status", "Site Name", "Date of Registration", 
        "Tax Validity", "Insurance Expiry", "Insurance Company", 
        "Insurance Policy Number", "PUC Expiry", "PUCC Number", 
        "Fitness Upto", "Registered At", "Source", "Last Synced", 
        "FASTag Company", "FASTag Balance", "FASTag Status", "Remark", 
        " Status", "Sold Date"
    ]
    
    # Helper function to clean date strings
    def clean_date(value):
        if not value:
            return ""
        # If it's already a string, remove the time part if it exists
        if isinstance(value, str):
            # Check if it's an ISO date string with time
            if 'T' in value:
                # Split at T and take only the date part
                return value.split('T')[0]
            # Try to parse as datetime
            try:
                dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                return dt.strftime('%Y-%m-%d')
            except:
                # If parsing fails, return as is
                return value
        # If it's a datetime object
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        return str(value)
    
    # Build CSV rows
    csv_rows = [",".join(f'"{h}"' for h in headers)]
    
    for v in vehicles:
        row = []
        for h in headers:
            # Get the raw value from the vehicle
            raw_value = None
            if h == "Registration Number":
                raw_value = v.get("registration_number", "")
            elif h == "Vehicle Type":
                raw_value = v.get("type", "")
            elif h == "Brand":
                raw_value = v.get("brand", "")
            elif h == "Model":
                raw_value = v.get("model", "")
            elif h == "Year":
                raw_value = v.get("year", "")
            elif h == "Chassis Number":
                raw_value = v.get("chassis_number", "")
            elif h == "Engine Number":
                raw_value = v.get("engine_number", "")
            elif h == "Color":
                raw_value = v.get("color", "")
            elif h == "Fuel Type":
                raw_value = v.get("fuel_type", "")
            elif h == "Average Mileage (km/l)":
                raw_value = v.get("average_kmpl", "")
            elif h == "Tank Capacity (L)":
                raw_value = v.get("tank_capacity_liters", "")
            elif h == "Seating Capacity":
                raw_value = v.get("seating_capacity", "")
            elif h == "Owner Name":
                raw_value = v.get("owner_name", "")
            elif h == "File Status":
                raw_value = "Complete" if v.get("file_status") else "Incomplete"
            elif h == "Site Name":
                raw_value = v.get("site_name", "")
            elif h == "Date of Registration":
                raw_value = clean_date(v.get("date_of_registration"))
            elif h == "Tax Validity":
                raw_value = clean_date(v.get("tax_upto"))
            elif h == "Insurance Expiry":
                raw_value = clean_date(v.get("insurance_expiry"))
            elif h == "Insurance Company":
                raw_value = v.get("insurance_company", "")
            elif h == "Insurance Policy Number":
                raw_value = v.get("insurance_policy_number", "")
            elif h == "PUC Expiry":
                raw_value = clean_date(v.get("puc_expiry"))
            elif h == "PUCC Number":
                raw_value = v.get("pucc_number", "")
            elif h == "Fitness Upto":
                raw_value = clean_date(v.get("fit_up_to"))
            elif h == "Registered At":
                raw_value = v.get("registered_at", "")
            elif h == "Source":
                raw_value = v.get("source", "Manual")
            elif h == "Last Synced":
                raw_value = clean_date(v.get("last_synced"))
            elif h == "FASTag Company":
                raw_value = v.get("fastag_company", "")
            elif h == "FASTag Balance":
                balance = v.get("fastag_balance")
                raw_value = f"₹{balance:,.2f}" if balance else ""
            elif h == "FASTag Status":
                raw_value = v.get("fastag_status", "")
            elif h == "Remark":
                raw_value = v.get("remark", "")
            elif h == "Sold Status":
                raw_value = "Yes" if v.get("sold") else "No"
            elif h == "Sold Date":
                raw_value = clean_date(v.get("sold_date"))
            
            # Clean the value for CSV
            if raw_value is None:
                val = ""
            elif isinstance(raw_value, bool):
                val = "Yes" if raw_value else "No"
            else:
                val = str(raw_value).replace('"', '""').replace(",", ";").replace("\n", " ")
            
            row.append(f'"{val}"')
        
        csv_rows.append(",".join(row))
    
    return {
        "csv_data": "\n".join(csv_rows),
        "filename": f"vehicles_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    }

@api_router.get("/vehicles/export/excel")
async def export_vehicles_excel(current_user: dict = Depends(get_current_user)):
    """Export all vehicles to a professional Excel file with all fields including document expiry dates"""
    try:
        vehicles = await db.vehicles.find({"is_deleted": False}, {"_id": 0}).to_list(1000)
        
        if not vehicles:
            raise HTTPException(status_code=404, detail="No vehicles found to export")
        
        # Helper function to clean date strings
        def clean_date(value):
            if not value:
                return ""
            if isinstance(value, str):
                if 'T' in value:
                    return value.split('T')[0]
                try:
                    dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                    return dt.strftime('%Y-%m-%d')
                except:
                    return value
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d')
            return str(value)
        
        # Get CURRENT documents for all vehicles
        vehicle_ids = [v["id"] for v in vehicles]
        
        # Fetch all current documents (only is_current = True)
        documents = await db.vehicle_documents.find(
            {
                "vehicle_id": {"$in": vehicle_ids},
                "is_current": True,
                "is_deleted": False
            },
            {"_id": 0}
        ).to_list(5000)
        
        # Group documents by vehicle_id
        docs_by_vehicle = {}
        for doc in documents:
            vehicle_id = doc["vehicle_id"]
            if vehicle_id not in docs_by_vehicle:
                docs_by_vehicle[vehicle_id] = {}
            
            doc_type = doc.get("document_type", "")
            docs_by_vehicle[vehicle_id][doc_type] = doc
        
        # Prepare data for Excel with document details
        excel_data = []
        for v in vehicles:
            vehicle_docs = docs_by_vehicle.get(v["id"], {})
            
            # Get current document expiry dates
            insurance_expiry = ""
            puc_expiry = ""
            fitness_expiry = ""
            tax_expiry = ""
            
            if "Insurance" in vehicle_docs:
                insurance_expiry = clean_date(vehicle_docs["Insurance"].get("expiry_date"))
            elif v.get("insurance_expiry"):
                # Fallback to vehicle field if no document found
                insurance_expiry = clean_date(v.get("insurance_expiry"))
            
            if "PUC" in vehicle_docs:
                puc_expiry = clean_date(vehicle_docs["PUC"].get("expiry_date"))
            elif v.get("puc_expiry"):
                puc_expiry = clean_date(v.get("puc_expiry"))
            
            if "Fitness" in vehicle_docs:
                fitness_expiry = clean_date(vehicle_docs["Fitness"].get("expiry_date"))
            elif v.get("fit_up_to"):
                fitness_expiry = clean_date(v.get("fit_up_to"))
            
            if "Tax" in vehicle_docs:
                tax_expiry = clean_date(vehicle_docs["Tax"].get("expiry_date"))
            elif v.get("tax_upto") and v.get("tax_upto") != "LIFETIME":
                tax_expiry = clean_date(v.get("tax_upto"))
            elif v.get("tax_upto") == "LIFETIME":
                tax_expiry = "LIFETIME"
            
            # Get document numbers
            insurance_policy = ""
            if "Insurance" in vehicle_docs:
                insurance_policy = vehicle_docs["Insurance"].get("policy_number", "")
            elif v.get("insurance_policy_number"):
                insurance_policy = v.get("insurance_policy_number", "")
            
            insurance_company = ""
            if "Insurance" in vehicle_docs:
                insurance_company = vehicle_docs["Insurance"].get("provider", "")
            elif v.get("insurance_company"):
                insurance_company = v.get("insurance_company", "")
            
            pucc_number = ""
            if "PUC" in vehicle_docs:
                pucc_number = vehicle_docs["PUC"].get("policy_number", "")
            elif v.get("pucc_number"):
                pucc_number = v.get("pucc_number", "")
            
            excel_data.append({
                "Registration Number": v.get("registration_number", ""),
                "Vehicle Type": v.get("type", ""),
                "Brand": v.get("brand", ""),
                "Model": v.get("model", ""),
                "Year": v.get("year", ""),
                "Chassis Number": v.get("chassis_number", ""),
                "Engine Number": v.get("engine_number", ""),
                "Color": v.get("color", ""),
                "Fuel Type": v.get("fuel_type", ""),
                "Seating Capacity": v.get("seating_capacity", ""),
                "Owner Name": v.get("owner_name", ""),
                "File Status": "Complete" if v.get("file_status") else "Incomplete",
                "Site Name": v.get("site_name", ""),
                "Date of Registration": clean_date(v.get("date_of_registration")),
                "Insurance Company": insurance_company,
                "Insurance Policy Number": insurance_policy,
                "Insurance Expiry": insurance_expiry,
                "PUCC Number": pucc_number,
                "PUC Expiry": puc_expiry,
                "Fitness Upto": fitness_expiry,
                "Tax Upto": tax_expiry,
                "Registered At": v.get("registered_at", ""),
                "Source": v.get("source", "Manual"),
                "Last Synced": clean_date(v.get("last_synced")),
                "FASTag Company": v.get("fastag_company", ""),
                "FASTag Balance": v.get("fastag_balance", ""),
                "FASTag Status": v.get("fastag_status", ""),
                "Remark": v.get("remark", ""),
                "Sold Status": "Yes" if v.get("sold") else "No",
                "Sold Date": clean_date(v.get("sold_date"))
            })
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        # Use openpyxl engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Vehicles', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Vehicles']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Color code expiry dates
            from openpyxl.styles import PatternFill, Font
            from openpyxl import load_workbook
            
            today = datetime.now(timezone.utc)
            today_str = today.strftime('%Y-%m-%d')
            
            # Define fills
            expired_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Red
            critical_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")  # Orange
            warning_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")  # Yellow
            valid_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Green
            
            # Find column indices for expiry dates
            col_map = {}
            for col_idx, cell in enumerate(worksheet[1], 1):
                if cell.value:
                    col_map[cell.value] = col_idx
            
            expiry_columns = ["Insurance Expiry", "PUC Expiry", "Fitness Upto", "Tax Upto"]
            expiry_col_indices = [col_map.get(col) for col in expiry_columns if col in col_map]
            
            # Apply color coding
            for row_idx in range(2, len(excel_data) + 2):
                for col_idx in expiry_col_indices:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    expiry_str = cell.value
                    
                    if expiry_str and expiry_str != "LIFETIME":
                        try:
                            expiry_date = datetime.strptime(str(expiry_str), '%Y-%m-%d').date()
                            today_date = today.date()
                            
                            if expiry_date < today_date:
                                cell.fill = expired_fill
                                cell.font = Font(color="FFFFFF", bold=True)
                            elif (expiry_date - today_date).days <= 7:
                                cell.fill = critical_fill
                                cell.font = Font(bold=True)
                            elif (expiry_date - today_date).days <= 30:
                                cell.fill = warning_fill
                        except Exception as e:
                            logger.warning(f"Could not parse expiry date: {expiry_str}")
            
            # Add summary sheet
            total_vehicles = len(vehicles)
            sold_vehicles = len([v for v in vehicles if v.get("sold")])
            active_vehicles = total_vehicles - sold_vehicles
            
            # Calculate document expiry statistics from current documents
            expired_insurance = 0
            expired_puc = 0
            expired_fitness = 0
            expired_tax = 0
            
            for doc in documents:
                if doc.get("document_type") == "Insurance":
                    expiry = clean_date(doc.get("expiry_date"))
                    if expiry and expiry < today_str:
                        expired_insurance += 1
                elif doc.get("document_type") == "PUC":
                    expiry = clean_date(doc.get("expiry_date"))
                    if expiry and expiry < today_str:
                        expired_puc += 1
                elif doc.get("document_type") == "Fitness":
                    expiry = clean_date(doc.get("expiry_date"))
                    if expiry and expiry < today_str:
                        expired_fitness += 1
                elif doc.get("document_type") == "Tax":
                    expiry = clean_date(doc.get("expiry_date"))
                    if expiry and expiry < today_str:
                        expired_tax += 1
            
            summary_data = {
                'Metric': [
                    'Total Vehicles',
                    'Active Vehicles',
                    'Sold Vehicles',
                    'Vehicles with Expired Insurance',
                    'Vehicles with Expired PUC',
                    'Vehicles with Expired Fitness',
                    'Vehicles with Expired Tax',
                    'Export Date'
                ],
                'Value': [
                    total_vehicles,
                    active_vehicles,
                    sold_vehicles,
                    expired_insurance,
                    expired_puc,
                    expired_fitness,
                    expired_tax,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Auto-adjust summary sheet columns
            summary_sheet = writer.sheets['Summary']
            summary_sheet.column_dimensions['A'].width = 35
            summary_sheet.column_dimensions['B'].width = 20
            
            # Add Documents Details sheet
            doc_data = []
            for doc in documents:
                vehicle = next((v for v in vehicles if v["id"] == doc["vehicle_id"]), {})
                doc_data.append({
                    "Registration Number": vehicle.get("registration_number", "Unknown"),
                    "Vehicle Model": f"{vehicle.get('brand', '')} {vehicle.get('model', '')}".strip(),
                    "Document Type": doc.get("document_type", ""),
                    "Policy Number": doc.get("policy_number", ""),
                    "Provider": doc.get("provider", ""),
                    "Issue Date": clean_date(doc.get("issue_date")),
                    "Expiry Date": clean_date(doc.get("expiry_date")),
                    "Status": doc.get("status", ""),
                    "Days Until Expiry": (
                        (datetime.strptime(clean_date(doc.get("expiry_date")), '%Y-%m-%d').date() - today.date()).days 
                        if clean_date(doc.get("expiry_date")) else "N/A"
                    ),
                    "Source": doc.get("source", "Manual")
                })
            
            if doc_data:
                doc_df = pd.DataFrame(doc_data)
                doc_df.to_excel(writer, sheet_name='Documents', index=False)
                
                # Auto-adjust documents sheet columns
                doc_sheet = writer.sheets['Documents']
                for column in doc_sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)
                    doc_sheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        filename = f"vehicles_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting vehicles to Excel: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export: {str(e)}")
    
@api_router.get("/vehicles/template/csv")
async def get_vehicle_import_template(current_user: dict = Depends(get_current_user)):
    """Get CSV template for vehicle import with sample data"""
    headers = [
        "registration_number", "type", "brand", "model", "year", "chassis_number",
        "engine_number", "color", "fuel_type", "average_kmpl", "tank_capacity_liters",
        "seating_capacity", "owner_name", "file_status", "site_name",
        "date_of_registration", "tax_upto", "remark"
    ]
    
    # Sample data rows
    sample_data = [
        ["MH-02-AB-1234", "Car", "Tata", "Nexon", "2024", "MABXX12345678901", "ENG123456", "White", "Diesel", "18.5", "44", "5", "John Doe", "Yes", "Mumbai HQ", "2024-01-15", "2025-03-31", "Company vehicle"],
        ["MH-12-CD-5678", "Truck", "Ashok Leyland", "Dost", "2023", "MABXX98765432101", "ENG654321", "Blue", "Diesel", "12.0", "90", "3", "Jane Smith", "No", "Pune Branch", "2023-06-20", "Tax valid till Dec 2025", "Delivery truck"],
        ["DL-01-EF-9012", "Van", "Maruti", "Eeco", "2022", "MABXX11223344556", "ENG112233", "Silver", "Petrol", "15.0", "40", "7", "Raj Kumar", "Yes", "Delhi Office", "2022-08-10", "2024-12-31", "Staff transport"]
    ]
    
    csv_rows = [",".join(headers)]
    for row in sample_data:
        csv_rows.append(",".join(row))
    
    return {
        "csv_data": "\n".join(csv_rows),
        "filename": "vehicle_import_template.csv",
        "instructions": {
            "type": "Car, Truck, Van, Bike, or Bus",
            "fuel_type": "Petrol, Diesel, Electric, CNG, or Hybrid",
            "file_status": "Yes or No",
            "date_of_registration": "YYYY-MM-DD format",
            "tax_upto": "Can be date (YYYY-MM-DD) or text description"
        }
    }

@api_router.post("/vehicles/import/csv")
async def import_vehicles_csv(data: dict, current_user: dict = Depends(get_current_user)):
    """Import vehicles from CSV data"""
    csv_data = data.get("csv_data", "")
    if not csv_data:
        raise HTTPException(status_code=400, detail="No CSV data provided")
    
    lines = csv_data.strip().split("\n")
    if len(lines) < 2:
        raise HTTPException(status_code=400, detail="CSV must have header row and at least one data row")
    
    headers = [h.strip().lower() for h in lines[0].split(",")]
    
    imported = 0
    errors = []
    
    for i, line in enumerate(lines[1:], start=2):
        try:
            values = line.split(",")
            if len(values) != len(headers):
                errors.append(f"Row {i}: Column count mismatch")
                continue
            
            row_data = dict(zip(headers, [v.strip() for v in values]))
            
            # Parse and validate
            vehicle_id = str(uuid.uuid4())
            vehicle_dict = {
                "id": vehicle_id,
                "registration_number": row_data.get("registration_number", ""),
                "type": row_data.get("type", "Car"),
                "brand": row_data.get("brand", ""),
                "model": row_data.get("model", ""),
                "year": int(row_data.get("year", 0)) if row_data.get("year") else None,
                "chassis_number": row_data.get("chassis_number") or None,
                "engine_number": row_data.get("engine_number") or None,
                "color": row_data.get("color") or None,
                "fuel_type": row_data.get("fuel_type", "Diesel"),
                "average_kmpl": float(row_data.get("average_kmpl", 10)),
                "tank_capacity_liters": float(row_data.get("tank_capacity_liters", 50)),
                "seating_capacity": int(row_data.get("seating_capacity", 0)) if row_data.get("seating_capacity") else None,
                "owner_name": row_data.get("owner_name") or None,
                "file_status": row_data.get("file_status", "").lower() in ["yes", "true", "1"],
                "site_name": row_data.get("site_name") or None,
                "date_of_registration": row_data.get("date_of_registration") or None,
                "tax_upto": row_data.get("tax_upto") or None,
                "remark": row_data.get("remark") or None,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user["user_id"],
                "is_deleted": False
            }
            
            if not vehicle_dict["registration_number"]:
                errors.append(f"Row {i}: Registration number is required")
                continue
            
            await db.vehicles.insert_one(vehicle_dict)
            imported += 1
            
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")
    
    return {
        "imported": imported,
        "errors": errors,
        "total_processed": len(lines) - 1
    }

@api_router.post("/fastag-passes")
async def create_fastag_pass(
    data: FastagPassCreate,
    current_user: dict = Depends(get_current_user)
):

    pass_id = str(uuid.uuid4())

    data_dict = data.model_dump()

    if not data_dict.get("balance_trips"):
        data_dict["balance_trips"] = data_dict["trips_allowed"]

    p = FastagPass(
        id=pass_id,
        **data_dict,
        created_by=current_user["user_id"]
    )

    pass_dict = p.model_dump()

    pass_dict["created_at"] = pass_dict["created_at"].isoformat()
    pass_dict["updated_at"] = pass_dict["updated_at"].isoformat()

    if isinstance(pass_dict["issue_date"], datetime):
        pass_dict["issue_date"] = pass_dict["issue_date"].isoformat()

    if isinstance(pass_dict["expiry_date"], datetime):
        pass_dict["expiry_date"] = pass_dict["expiry_date"].isoformat()

    await db.fastag_passes.insert_one(pass_dict)

    return serialize_doc(pass_dict)

@api_router.delete("/fastag-passes/{pass_id}")
async def delete_fastag_pass(
    pass_id: str,
    current_user: dict = Depends(get_current_user)
):

    result = await db.fastag_passes.update_one(
        {"id": pass_id},
        {
            "$set": {
                "is_deleted": True,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pass not found")

    return {"message": "FASTag pass deleted"}    
    
@api_router.get("/vehicles/{vehicle_id}/fastag-passes")
async def get_fastag_passes(vehicle_id: str, current_user: dict = Depends(get_current_user)):

    passes = await db.fastag_passes.find(
        {"vehicle_id": vehicle_id, "is_deleted": False},
        {"_id": 0}
    ).to_list(100)

    now = datetime.now(timezone.utc)

    for p in passes:
        expiry = datetime.fromisoformat(p["expiry_date"])
        if expiry < now:
            p["status"] = "Inactive"

    return {"data": passes}

@api_router.put("/fastag-passes/{pass_id}/use-trip")
async def use_fastag_trip(pass_id: str, current_user: dict = Depends(get_current_user)):

    await db.fastag_passes.update_one(
    {"id": pass_id},
    {"$inc": {"balance_trips": -1}}
)

    return {"message": "Trip used"}    

@api_router.put("/fastag-passes/{pass_id}")
async def update_fastag_pass(
    pass_id: str,
    data: FastagPassCreate,
    current_user: dict = Depends(get_current_user)
):

    update = data.model_dump()

    if isinstance(update["issue_date"], datetime):
        update["issue_date"] = update["issue_date"].isoformat()

    if isinstance(update["expiry_date"], datetime):
        update["expiry_date"] = update["expiry_date"].isoformat()

    update["updated_at"] = datetime.now(timezone.utc).isoformat()

    await db.fastag_passes.update_one(
        {"id": pass_id},
        {"$set": update}
    )

    return {"message": "Pass updated"}

@api_router.post("/surepass/fetch-vehicle")
async def fetch_vehicle_from_surepass(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Fetch vehicle details from Surepass API
    First checks if vehicle exists and if documents need update
    """
    registration_number = data.get("registration_number", "").upper().strip()
    force_refresh = data.get("force_refresh", False)
    
    if not registration_number:
        raise HTTPException(status_code=400, detail="Registration number is required")
    
    # FIX: Use case-insensitive regex search to find existing vehicle
    existing_vehicle = await db.vehicles.find_one({
        "registration_number": {"$regex": f"^{registration_number}$", "$options": "i"},
        "is_deleted": False
    })
    
    logger.info(f"Checking for vehicle {registration_number}: {'Found' if existing_vehicle else 'Not found'}")
    
    # If vehicle exists and not forcing refresh, check if documents need update
    if existing_vehicle and not force_refresh:
        # Check document status
        insurance_status = surepass_service.check_document_status(
            existing_vehicle.get("insurance_expiry")
        )
        puc_status = surepass_service.check_document_status(
            existing_vehicle.get("puc_expiry")
        )
        
        # If both documents are valid (>30 days), no need to fetch
        if not insurance_status["needs_update"] and not puc_status["needs_update"]:
            return {
                "exists": True,
                "needs_update": False,
                "vehicle": serialize_doc(existing_vehicle),
                "document_status": {
                    "insurance": insurance_status,
                    "puc": puc_status
                },
                "message": "Vehicle already exists with valid documents. No update needed."
            }
        
        # If documents need update, proceed with API call
        existing_id = existing_vehicle["id"]
        logger.info(f"Vehicle exists but needs update. ID: {existing_id}")
    else:
        existing_id = None
        if existing_vehicle and force_refresh:
            existing_id = existing_vehicle["id"]
            logger.info(f"Force refresh for existing vehicle. ID: {existing_id}")
    
    # Call Surepass API
    result = await surepass_service.fetch_vehicle_details(registration_number)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result.get("error", "Failed to fetch vehicle details"))
    
    # Parse the data
    parsed_data = surepass_service.parse_vehicle_data(result)
    
    # Log the API call
    log_entry = {
        "id": str(uuid.uuid4()),
        "registration_number": registration_number,
        "request_timestamp": datetime.now(timezone.utc).isoformat(),
        "response_data": result.get("raw_response", {}),
        "insurance_expiry": parsed_data.get("insurance_expiry").isoformat() if parsed_data.get("insurance_expiry") else None,
        "puc_expiry": parsed_data.get("puc_expiry").isoformat() if parsed_data.get("puc_expiry") else None,
        "fit_up_to": parsed_data.get("fit_up_to").isoformat() if parsed_data.get("fit_up_to") else None,
        "tax_upto": parsed_data.get("tax_upto").isoformat() if parsed_data.get("tax_upto") else None,
        "is_successful": True,
        "created_by": current_user["user_id"]
    }
    await db.rc_verification_logs.insert_one(log_entry)
    
    # Return based on whether vehicle exists
    if existing_id:
        return {
            "exists": True,
            "needs_update": True,
            "vehicle_data": parsed_data,
            "vehicle_id": existing_id,
            "message": "Vehicle exists but documents need update. Please review and save."
        }
    
    # New vehicle
    return {
        "exists": False,
        "needs_update": True,
        "vehicle_data": parsed_data,
        "message": "Vehicle details fetched successfully. Please review and save."
    }

@api_router.post("/vehicles/{vehicle_id}/sync-documents")
async def sync_vehicle_documents(
    vehicle_id: str,
    current_user: dict = Depends(get_current_user)
):
    vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")

    registration_number = vehicle["registration_number"]
    today = datetime.now(timezone.utc)

    def is_expired(date_value):
        if not date_value:
            return False
        try:
            if isinstance(date_value, str):
                dt = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
            else:
                dt = date_value
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt < today
        except:
            return False

    def is_valid(date_value):
        """Check if document is valid (not expired)"""
        return not is_expired(date_value)

    # Check all 5 document types on the vehicle record
    needs_sync = (
        is_expired(vehicle.get("insurance_expiry")) or
        is_expired(vehicle.get("puc_expiry")) or
        is_expired(vehicle.get("fit_up_to")) or
        (vehicle.get("tax_upto") and vehicle.get("tax_upto") != "LIFETIME" and is_expired(vehicle.get("tax_upto")))
    )

    if not needs_sync:
        return {
            "success": True,
            "vehicle_id": vehicle_id,
            "registration_number": registration_number,
            "documents_created": [],
            "message": "All documents are valid. No sync needed."
        }

    # Fetch latest data from Surepass
    result = await surepass_service.fetch_vehicle_details(registration_number)

    if not result["success"]:
        raise HTTPException(status_code=400, detail=result.get("error", "Failed to fetch vehicle details"))

    parsed_data = surepass_service.parse_vehicle_data(result)

    def parse_date(date_value):
        if not date_value:
            return None
        if isinstance(date_value, datetime):
            return date_value.isoformat()
        if isinstance(date_value, str):
            try:
                dt = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
                return dt.isoformat()
            except (ValueError, AttributeError):
                return date_value
        return date_value

    def determine_status(expiry_value):
        if not expiry_value:
            return "Active"
        try:
            if isinstance(expiry_value, str):
                dt = datetime.fromisoformat(expiry_value.replace('Z', '+00:00'))
            else:
                dt = expiry_value
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return "Expired" if dt < today else "Active"
        except:
            return "Active"

    async def version_document(vehicle_id, doc_type, new_doc_data):
        existing = await db.vehicle_documents.find_one({
            "vehicle_id": vehicle_id,
            "document_type": doc_type,
            "is_current": True,
            "is_deleted": False
        })

        if existing:
            await db.vehicle_documents.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "is_current": False,
                    "is_past_document": True,
                    "updated_at": today.isoformat()
                }}
            )
            new_doc_data["previous_version_id"] = existing["id"]
            new_doc_data["version"] = existing.get("version", 0) + 1
        else:
            new_doc_data["version"] = 1

        new_doc_data["id"] = str(uuid.uuid4())
        new_doc_data["is_current"] = True
        new_doc_data["is_past_document"] = False
        new_doc_data["created_at"] = today.isoformat()
        new_doc_data["updated_at"] = today.isoformat()
        new_doc_data["created_by"] = current_user["user_id"]
        new_doc_data["is_deleted"] = False

        await db.vehicle_documents.insert_one(new_doc_data)
        return new_doc_data["id"]

    documents_updated = []
    vehicle_update = {
        "last_synced": today.isoformat(),
        "updated_at": today.isoformat(),
    }

    # ✅ Insurance — sync if expired in vehicle record AND valid in Surepass data
    if parsed_data.get("insurance_expiry"):
        vehicle_insurance_expired = is_expired(vehicle.get("insurance_expiry"))
        surepass_insurance_valid = is_valid(parsed_data.get("insurance_expiry"))
        
        if vehicle_insurance_expired and surepass_insurance_valid:
            existing_insurance = await db.vehicle_documents.find_one({
                "vehicle_id": vehicle_id,
                "document_type": "Insurance",
                "is_current": True,
                "is_deleted": False
            })
            doc_expired = existing_insurance is None or is_expired(existing_insurance.get("expiry_date"))
            
            if doc_expired:
                insurance_doc = {
                    "vehicle_id": vehicle_id,
                    "document_type": "Insurance",
                    "policy_number": parsed_data.get("insurance_policy_number") or f"INS-{registration_number}",
                    "provider": parsed_data.get("insurance_company") or "Unknown",
                    "issue_date": parse_date(parsed_data.get("date_of_registration")) or today.isoformat(),
                    "expiry_date": parse_date(parsed_data.get("insurance_expiry")),
                    "status": "Active",
                    "source": "surepass",
                }
                await version_document(vehicle_id, "Insurance", insurance_doc)
                documents_updated.append("Insurance")
                vehicle_update["insurance_expiry"] = parse_date(parsed_data.get("insurance_expiry"))
                vehicle_update["insurance_company"] = parsed_data.get("insurance_company")
                vehicle_update["insurance_policy_number"] = parsed_data.get("insurance_policy_number")
                logger.info(f"Updated Insurance for {registration_number} from expired to valid")

    # ✅ PUC — sync if expired in vehicle record AND valid in Surepass data
    if parsed_data.get("puc_expiry"):
        vehicle_puc_expired = is_expired(vehicle.get("puc_expiry"))
        surepass_puc_valid = is_valid(parsed_data.get("puc_expiry"))
        
        if vehicle_puc_expired and surepass_puc_valid:
            existing_puc = await db.vehicle_documents.find_one({
                "vehicle_id": vehicle_id,
                "document_type": "PUC",
                "is_current": True,
                "is_deleted": False
            })
            doc_expired = existing_puc is None or is_expired(existing_puc.get("expiry_date"))
            
            if doc_expired:
                puc_doc = {
                    "vehicle_id": vehicle_id,
                    "document_type": "PUC",
                    "policy_number": parsed_data.get("pucc_number") or f"PUC-{registration_number}",
                    "provider": parsed_data.get("registered_at") or "RTO",
                    "issue_date": parse_date(parsed_data.get("date_of_registration")) or today.isoformat(),
                    "expiry_date": parse_date(parsed_data.get("puc_expiry")),
                    "status": "Active",
                    "source": "surepass",
                }
                await version_document(vehicle_id, "PUC", puc_doc)
                documents_updated.append("PUC")
                vehicle_update["puc_expiry"] = parse_date(parsed_data.get("puc_expiry"))
                vehicle_update["pucc_number"] = parsed_data.get("pucc_number")
                logger.info(f"Updated PUC for {registration_number} from expired to valid")

    # ✅ Fitness — sync if expired in vehicle record AND valid in Surepass data
    if parsed_data.get("fit_up_to"):
        vehicle_fitness_expired = is_expired(vehicle.get("fit_up_to"))
        surepass_fitness_valid = is_valid(parsed_data.get("fit_up_to"))
        
        if vehicle_fitness_expired and surepass_fitness_valid:
            existing_fitness = await db.vehicle_documents.find_one({
                "vehicle_id": vehicle_id,
                "document_type": "Fitness",
                "is_current": True,
                "is_deleted": False
            })
            doc_expired = existing_fitness is None or is_expired(existing_fitness.get("expiry_date"))
            
            if doc_expired:
                fitness_doc = {
                    "vehicle_id": vehicle_id,
                    "document_type": "Fitness",
                    "policy_number": f"FIT-{registration_number}",
                    "provider": parsed_data.get("registered_at") or "RTO",
                    "issue_date": parse_date(parsed_data.get("date_of_registration")) or today.isoformat(),
                    "expiry_date": parse_date(parsed_data.get("fit_up_to")),
                    "status": "Active",
                    "source": "surepass",
                }
                await version_document(vehicle_id, "Fitness", fitness_doc)
                documents_updated.append("Fitness")
                vehicle_update["fit_up_to"] = parse_date(parsed_data.get("fit_up_to"))
                logger.info(f"Updated Fitness for {registration_number} from expired to valid")

    # ✅ NEW: Tax Document — sync if expired in vehicle record AND valid in Surepass data
    if parsed_data.get("tax_upto"):
        # Check if tax_upto is not LIFETIME (which means it's a date)
        vehicle_tax = vehicle.get("tax_upto")
        if vehicle_tax and vehicle_tax != "LIFETIME":
            vehicle_tax_expired = is_expired(vehicle_tax)
            surepass_tax_valid = is_valid(parsed_data.get("tax_upto"))
            
            if vehicle_tax_expired and surepass_tax_valid:
                existing_tax = await db.vehicle_documents.find_one({
                    "vehicle_id": vehicle_id,
                    "document_type": "Tax",
                    "is_current": True,
                    "is_deleted": False
                })
                doc_expired = existing_tax is None or is_expired(existing_tax.get("expiry_date"))
                
                if doc_expired:
                    tax_doc = {
                        "vehicle_id": vehicle_id,
                        "document_type": "Tax",
                        "policy_number": f"TAX-{registration_number}",
                        "provider": parsed_data.get("registered_at") or "RTO",
                        "issue_date": parse_date(parsed_data.get("date_of_registration")) or today.isoformat(),
                        "expiry_date": parse_date(parsed_data.get("tax_upto")),
                        "status": "Active",
                        "source": "surepass",
                    }
                    await version_document(vehicle_id, "Tax", tax_doc)
                    documents_updated.append("Tax")
                    vehicle_update["tax_upto"] = parse_date(parsed_data.get("tax_upto"))
                    logger.info(f"Updated Tax for {registration_number} from expired to valid")

    # ✅ Update vehicle record only with fields that were actually synced
    if documents_updated:
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": vehicle_update}
        )

    return {
        "success": True,
        "vehicle_id": vehicle_id,
        "registration_number": registration_number,
        "documents_created": documents_updated,
        "message": (
            f"Updated {len(documents_updated)} document(s) from expired to valid: {', '.join(documents_updated)}"
            if documents_updated
            else "No expired documents had valid replacements from Surepass"
        )
    }

@api_router.post("/vehicles/from-surepass")
async def create_vehicle_from_surepass(
    vehicle_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Create or update vehicle from Surepass data
    ALWAYS creates insurance, PUC, Fitness, and Tax documents regardless of expiry status
    """
    registration_number = vehicle_data.get("registration_number")
    
    logger.info(f"=== CREATING VEHICLE FROM SUREPASS: {registration_number} ===")
    logger.info(f"Vehicle data received: {json.dumps(vehicle_data, default=str, indent=2)}")
    
    # Check if vehicle already exists
    existing = await db.vehicles.find_one({
        "registration_number": registration_number,
        "is_deleted": False
    })
    
    logger.info(f"Vehicle exists: {bool(existing)}")
    
    # Prepare vehicle document
    vehicle_id = existing["id"] if existing else str(uuid.uuid4())
    logger.info(f"Vehicle ID: {vehicle_id}")
    
    # Helper function to handle date parsing
    def parse_date(date_value):
        """Convert date string or datetime to ISO format string"""
        if not date_value:
            return None
        if isinstance(date_value, datetime):
            return date_value.isoformat()
        if isinstance(date_value, str):
            try:
                # Handle format like "2027-01-24T00:00:00"
                if 'T' in date_value:
                    dt = datetime.fromisoformat(date_value)
                    return dt.isoformat()
                # Handle simple date format
                else:
                    dt = datetime.fromisoformat(date_value)
                    return dt.isoformat()
            except (ValueError, AttributeError) as e:
                logger.warning(f"Date parsing error for {date_value}: {e}")
                return date_value
        return date_value
    
    # Map Surepass fields to our Vehicle model
    vehicle_dict = {
        "id": vehicle_id,
        "registration_number": registration_number,
        "type": vehicle_data.get("type", "Car"),
        "brand": vehicle_data.get("brand", ""),
        "model": vehicle_data.get("model", ""),
        "year": int(vehicle_data.get("year")) if vehicle_data.get("year") else None,
        "chassis_number": vehicle_data.get("chassis_number"),
        "engine_number": vehicle_data.get("engine_number"),
        "color": vehicle_data.get("color"),
        "fuel_type": vehicle_data.get("fuel_type", "Diesel"),
        "seating_capacity": int(vehicle_data.get("seating_capacity")) if vehicle_data.get("seating_capacity") else None,
        "owner_name": vehicle_data.get("owner_name"),
        "date_of_registration": parse_date(vehicle_data.get("date_of_registration")),
        "insurance_expiry": parse_date(vehicle_data.get("insurance_expiry")),
        "puc_expiry": parse_date(vehicle_data.get("puc_expiry")),
        "fit_up_to": parse_date(vehicle_data.get("fit_up_to")),
        "tax_upto": parse_date(vehicle_data.get("tax_upto")),
        "insurance_company": vehicle_data.get("insurance_company"),
        "insurance_policy_number": vehicle_data.get("insurance_policy_number"),
        "pucc_number": vehicle_data.get("pucc_number"),
        "registered_at": vehicle_data.get("registered_at"),
        "source": "surepass",
        "last_synced": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["user_id"],
        "is_deleted": False
    }
    
    # Add manual fields if they exist
    if vehicle_data.get("site_name"):
        vehicle_dict["site_name"] = vehicle_data["site_name"]
    if vehicle_data.get("remark"):
        vehicle_dict["remark"] = vehicle_data["remark"]
    if vehicle_data.get("file_status") is not None:
        vehicle_dict["file_status"] = vehicle_data["file_status"]
    
    logger.info(f"Vehicle dict prepared: {json.dumps(vehicle_dict, default=str, indent=2)}")
    
    if existing:
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": vehicle_dict}
        )
        message = "Vehicle updated successfully from Surepass"
        logger.info(f"Updated vehicle {registration_number}")
    else:
        vehicle_dict["created_at"] = datetime.now(timezone.utc).isoformat()
        vehicle_dict["created_by"] = current_user["user_id"]
        await db.vehicles.insert_one(vehicle_dict)
        message = "Vehicle created successfully from Surepass"
        logger.info(f"Created vehicle {registration_number}")
    
    # ============ ALWAYS CREATE INSURANCE DOCUMENT ============
    if vehicle_data.get("insurance_expiry"):
        logger.info(f"Creating Insurance document for {registration_number}")
        
        # Parse expiry date for status determination
        insurance_status = "Active"
        insurance_expiry = vehicle_data.get("insurance_expiry")
        try:
            if isinstance(insurance_expiry, str):
                if 'T' in insurance_expiry:
                    expiry_date = datetime.fromisoformat(insurance_expiry)
                else:
                    expiry_date = datetime.fromisoformat(insurance_expiry)
            else:
                expiry_date = insurance_expiry
            
            if expiry_date.tzinfo is None:
                expiry_date = expiry_date.replace(tzinfo=timezone.utc)
            
            if expiry_date < datetime.now(timezone.utc):
                insurance_status = "Expired"
        except Exception as e:
            logger.error(f"Error parsing insurance expiry: {e}")
            expiry_date = None
        
        insurance_data = {
            "id": str(uuid.uuid4()),
            "vehicle_id": vehicle_id,
            "document_type": "Insurance",
            "policy_number": vehicle_data.get("insurance_policy_number", f"INS-{registration_number}"),
            "provider": vehicle_data.get("insurance_company", "Unknown"),
            "issue_date": parse_date(vehicle_data.get("date_of_registration")) or datetime.now(timezone.utc).isoformat(),
            "expiry_date": parse_date(vehicle_data.get("insurance_expiry")),
            "status": insurance_status,
            "version": 1,
            "is_current": True,
            "source": "surepass",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "is_deleted": False
        }
        
        logger.info(f"Insurance data prepared: {json.dumps(insurance_data, default=str)}")
        
        existing_insurance = await db.vehicle_documents.find_one({
            "vehicle_id": vehicle_id,
            "document_type": "Insurance",
            "is_current": True
        })
        
        if existing_insurance:
            await db.vehicle_documents.update_many(
                {"vehicle_id": vehicle_id, "document_type": "Insurance", "is_current": True},
                {"$set": {"is_current": False}}
            )
            insurance_data["previous_version_id"] = existing_insurance["id"]
            insurance_data["version"] = existing_insurance.get("version", 0) + 1
        
        result = await db.vehicle_documents.insert_one(insurance_data)
        logger.info(f"Created Insurance document with ID: {insurance_data['id']}, insert result: {result.inserted_id}")
    else:
        logger.info(f"No insurance expiry data found for {registration_number}")
    
    # ============ ALWAYS CREATE PUC DOCUMENT ============
    if vehicle_data.get("puc_expiry"):
        logger.info(f"Creating PUC document for {registration_number}")
        
        puc_status = "Active"
        puc_expiry = vehicle_data.get("puc_expiry")
        try:
            if isinstance(puc_expiry, str):
                if 'T' in puc_expiry:
                    expiry_date = datetime.fromisoformat(puc_expiry)
                else:
                    expiry_date = datetime.fromisoformat(puc_expiry)
            else:
                expiry_date = puc_expiry
            
            if expiry_date.tzinfo is None:
                expiry_date = expiry_date.replace(tzinfo=timezone.utc)
            
            if expiry_date < datetime.now(timezone.utc):
                puc_status = "Expired"
        except Exception as e:
            logger.error(f"Error parsing PUC expiry: {e}")
            expiry_date = None
        
        puc_data = {
            "id": str(uuid.uuid4()),
            "vehicle_id": vehicle_id,
            "document_type": "PUC",
            "policy_number": vehicle_data.get("pucc_number", f"PUC-{registration_number}"),
            "provider": vehicle_data.get("registered_at", "RTO"),
            "issue_date": parse_date(vehicle_data.get("date_of_registration")) or datetime.now(timezone.utc).isoformat(),
            "expiry_date": parse_date(vehicle_data.get("puc_expiry")),
            "status": puc_status,
            "version": 1,
            "is_current": True,
            "source": "surepass",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "is_deleted": False
        }
        
        logger.info(f"PUC data prepared: {json.dumps(puc_data, default=str)}")
        
        existing_puc = await db.vehicle_documents.find_one({
            "vehicle_id": vehicle_id,
            "document_type": "PUC",
            "is_current": True
        })
        
        if existing_puc:
            await db.vehicle_documents.update_many(
                {"vehicle_id": vehicle_id, "document_type": "PUC", "is_current": True},
                {"$set": {"is_current": False}}
            )
            puc_data["previous_version_id"] = existing_puc["id"]
            puc_data["version"] = existing_puc.get("version", 0) + 1
        
        result = await db.vehicle_documents.insert_one(puc_data)
        logger.info(f"Created PUC document with ID: {puc_data['id']}, insert result: {result.inserted_id}")
    else:
        logger.info(f"No PUC expiry data found for {registration_number}")    
    
    # ============ ALWAYS CREATE FITNESS DOCUMENT ============
    if vehicle_data.get("fit_up_to"):
        logger.info(f"Creating Fitness document for {registration_number}")
        
        fitness_status = "Active"
        fitness_expiry = vehicle_data.get("fit_up_to")
        try:
            if isinstance(fitness_expiry, str):
                if 'T' in fitness_expiry:
                    expiry_date = datetime.fromisoformat(fitness_expiry)
                else:
                    expiry_date = datetime.fromisoformat(fitness_expiry)
            else:
                expiry_date = fitness_expiry
            
            if expiry_date.tzinfo is None:
                expiry_date = expiry_date.replace(tzinfo=timezone.utc)
            
            if expiry_date < datetime.now(timezone.utc):
                fitness_status = "Expired"
        except Exception as e:
            logger.error(f"Error parsing fitness expiry: {e}")
            expiry_date = None
        
        fitness_data = {
            "id": str(uuid.uuid4()),
            "vehicle_id": vehicle_id,
            "document_type": "Fitness",
            "policy_number": f"FIT-{registration_number}",
            "provider": vehicle_data.get("registered_at", "RTO"),
            "issue_date": parse_date(vehicle_data.get("date_of_registration")) or datetime.now(timezone.utc).isoformat(),
            "expiry_date": parse_date(vehicle_data.get("fit_up_to")),
            "status": fitness_status,
            "version": 1,
            "is_current": True,
            "source": "surepass",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "is_deleted": False
        }
        
        logger.info(f"Fitness data prepared: {json.dumps(fitness_data, default=str)}")
        
        existing_fitness = await db.vehicle_documents.find_one({
            "vehicle_id": vehicle_id,
            "document_type": "Fitness",
            "is_current": True
        })
        
        if existing_fitness:
            await db.vehicle_documents.update_many(
                {"vehicle_id": vehicle_id, "document_type": "Fitness", "is_current": True},
                {"$set": {"is_current": False}}
            )
            fitness_data["previous_version_id"] = existing_fitness["id"]
            fitness_data["version"] = existing_fitness.get("version", 0) + 1
        
        result = await db.vehicle_documents.insert_one(fitness_data)
        logger.info(f"Created Fitness document with ID: {fitness_data['id']}, insert result: {result.inserted_id}")
    else:
        logger.info(f"No fitness expiry data found for {registration_number}")
    
    # ============ ALWAYS CREATE TAX DOCUMENT (if not LIFETIME) ============
    if vehicle_data.get("tax_upto") and vehicle_data.get("tax_upto") != "LIFETIME":
        logger.info(f"Creating Tax document for {registration_number}")
        
        tax_status = "Active"
        tax_expiry = vehicle_data.get("tax_upto")
        try:
            if isinstance(tax_expiry, str):
                if 'T' in tax_expiry:
                    expiry_date = datetime.fromisoformat(tax_expiry)
                else:
                    expiry_date = datetime.fromisoformat(tax_expiry)
            else:
                expiry_date = tax_expiry
            
            if expiry_date.tzinfo is None:
                expiry_date = expiry_date.replace(tzinfo=timezone.utc)
            
            if expiry_date < datetime.now(timezone.utc):
                tax_status = "Expired"
        except Exception as e:
            logger.error(f"Error parsing tax expiry: {e}")
            expiry_date = None
        
        tax_data = {
            "id": str(uuid.uuid4()),
            "vehicle_id": vehicle_id,
            "document_type": "Tax",
            "policy_number": f"TAX-{registration_number}",
            "provider": vehicle_data.get("registered_at", "RTO"),
            "issue_date": parse_date(vehicle_data.get("date_of_registration")) or datetime.now(timezone.utc).isoformat(),
            "expiry_date": parse_date(vehicle_data.get("tax_upto")),
            "status": tax_status,
            "version": 1,
            "is_current": True,
            "source": "surepass",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "is_deleted": False
        }
        
        logger.info(f"Tax data prepared: {json.dumps(tax_data, default=str)}")
        
        existing_tax = await db.vehicle_documents.find_one({
            "vehicle_id": vehicle_id,
            "document_type": "Tax",
            "is_current": True
        })
        
        if existing_tax:
            await db.vehicle_documents.update_many(
                {"vehicle_id": vehicle_id, "document_type": "Tax", "is_current": True},
                {"$set": {"is_current": False}}
            )
            tax_data["previous_version_id"] = existing_tax["id"]
            tax_data["version"] = existing_tax.get("version", 0) + 1
        
        result = await db.vehicle_documents.insert_one(tax_data)
        logger.info(f"Created Tax document with ID: {tax_data['id']}, insert result: {result.inserted_id}")
    elif vehicle_data.get("tax_upto") == "LIFETIME":
        logger.info(f"Vehicle {registration_number} has lifetime tax, skipping Tax document creation")
    else:
        logger.info(f"No tax expiry data found for {registration_number}")
    
    logger.info(f"=== FINISHED CREATING VEHICLE {registration_number} ===")
    
    return {
        "success": True,
        "vehicle_id": vehicle_id,
        "message": message
    }

@api_router.get("/debug/vehicle-documents/{vehicle_id}")
async def debug_vehicle_documents(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    """Debug endpoint to check documents for a vehicle"""
    # Find the vehicle first by ID or registration number
    vehicle = await db.vehicles.find_one({
        "$or": [
            {"id": vehicle_id},
            {"registration_number": vehicle_id.upper()}
        ],
        "is_deleted": False
    })
    
    if not vehicle:
        return {"error": "Vehicle not found", "searched_for": vehicle_id}
    
    # Get all documents for this vehicle (including deleted ones for debugging)
    documents = await db.vehicle_documents.find(
        {"vehicle_id": vehicle["id"]},
        {"_id": 0}
    ).to_list(100)
    
    # Also check if there's any issue with the document fields
    doc_fields = []
    for doc in documents:
        doc_fields.append({
            "id": doc.get("id"),
            "document_type": doc.get("document_type"),
            "is_current": doc.get("is_current"),
            "is_deleted": doc.get("is_deleted", False)
        })
    
    return {
        "vehicle": {
            "id": vehicle["id"],
            "registration_number": vehicle["registration_number"]
        },
        "document_count": len(documents),
        "documents": documents,
        "document_fields_summary": doc_fields
    }

# ==================== VEHICLE DOCUMENT ROUTES (VERSIONED) ====================

@api_router.post("/vehicle-documents")
async def create_vehicle_document(
    document_data: VehicleDocumentCreate,
    current_user: dict = Depends(get_current_user)
):
    try:
        vehicle = await db.vehicles.find_one({"id": document_data.vehicle_id, "is_deleted": False})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")

        # Check for existing current version
        existing_doc = await db.vehicle_documents.find_one({
            "vehicle_id": document_data.vehicle_id,
            "document_type": document_data.document_type,
            "is_current": True,
            "is_deleted": False
        })

        # If is_current is explicitly False, this is a past document
        is_past_document = document_data.is_current == False

        if is_past_document:
            # Validate: past document expiry must be before the current doc's issue date
            if existing_doc:
                current_issue_date = existing_doc.get("issue_date")
                if current_issue_date:
                    if isinstance(current_issue_date, str):
                        current_issue_dt = datetime.fromisoformat(current_issue_date.replace('Z', '+00:00'))
                    else:
                        current_issue_dt = current_issue_date

                    if current_issue_dt.tzinfo is None:
                        current_issue_dt = current_issue_dt.replace(tzinfo=timezone.utc)

                    expiry_date = document_data.expiry_date
                    if isinstance(expiry_date, str):
                        expiry_dt = datetime.fromisoformat(expiry_date.replace('Z', '+00:00'))
                    else:
                        expiry_dt = expiry_date

                    if expiry_dt.tzinfo is None:
                        expiry_dt = expiry_dt.replace(tzinfo=timezone.utc)

                    if expiry_dt >= current_issue_dt:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Past document expiry date ({expiry_dt.strftime('%Y-%m-%d')}) must be before the current document's issue date ({current_issue_dt.strftime('%Y-%m-%d')})"
                        )

        document_id = str(uuid.uuid4())

        issue_date = document_data.issue_date
        if isinstance(issue_date, datetime):
            issue_date = issue_date.isoformat()

        expiry_date = document_data.expiry_date
        if isinstance(expiry_date, datetime):
            expiry_date = expiry_date.isoformat()

        document_dict = {
            "id": document_id,
            "vehicle_id": document_data.vehicle_id,
            "document_type": document_data.document_type,
            "custom_document_name": document_data.custom_document_name,
            "policy_number": document_data.policy_number,
            "provider": document_data.provider,
            "phone_number": document_data.phone_number,
            "issue_date": issue_date,
            "expiry_date": expiry_date,
            "premium": document_data.premium,
            "coverage": document_data.coverage,
            "status": document_data.status,
            "file_url": document_data.file_url,
            "file_public_id": document_data.file_public_id,
            "file_type": document_data.file_type,
            "file_uploaded_at": document_data.file_uploaded_at.isoformat() if document_data.file_uploaded_at else None,
            "file_uploaded_by": document_data.file_uploaded_by,
            "version": (existing_doc.get("version", 0) + 1) if existing_doc and not is_past_document else 1,
            "is_current": False if is_past_document else True,
            "is_past_document": is_past_document,
            "previous_version_id": existing_doc["id"] if existing_doc and not is_past_document else None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user["user_id"],
            "updated_by": current_user["user_id"],
            "is_deleted": False
        }

        # Only mark existing doc as not current if this is NOT a past document
        if existing_doc and not is_past_document:
            await db.vehicle_documents.update_many(
                {
                    "vehicle_id": document_data.vehicle_id,
                    "document_type": document_data.document_type,
                    "is_current": True
                },
                {"$set": {"is_current": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
            )

        await db.vehicle_documents.insert_one(document_dict)

        # 🔥 NEW: Update vehicle's main fields if this is a current document
        if not is_past_document and document_data.is_current != False:
            vehicle_update = {}
            doc_type = document_data.document_type
            
            if doc_type == "Insurance":
                vehicle_update["insurance_expiry"] = expiry_date
                vehicle_update["insurance_company"] = document_data.provider
                vehicle_update["insurance_policy_number"] = document_data.policy_number
            elif doc_type == "PUC":
                vehicle_update["puc_expiry"] = expiry_date
                vehicle_update["pucc_number"] = document_data.policy_number
            elif doc_type == "Fitness":
                vehicle_update["fit_up_to"] = expiry_date
            elif doc_type == "Tax":
                vehicle_update["tax_upto"] = expiry_date
            
            if vehicle_update:
                vehicle_update["updated_at"] = datetime.now(timezone.utc).isoformat()
                vehicle_update["updated_by"] = current_user["user_id"]
                await db.vehicles.update_one(
                    {"id": document_data.vehicle_id},
                    {"$set": vehicle_update}
                )
                logger.info(f"Updated vehicle {document_data.vehicle_id} {doc_type} expiry to {expiry_date}")

        return {
            "success": True,
            "document_id": document_id,
            "version": document_dict["version"],
            "is_past_document": is_past_document,
            "message": "Past document added successfully" if is_past_document else "Document created successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating vehicle document: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to create document: {str(e)}")

@api_router.get("/vehicles/{vehicle_id}/document-status")
async def check_vehicle_document_status(
    vehicle_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Check if vehicle documents need update based on expiry
    """
    vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Get current documents
    documents = await db.vehicle_documents.find({
        "vehicle_id": vehicle_id,
        "is_current": True,
        "is_deleted": False
    }).to_list(10)
    
    document_status = {}
    needs_update = False
    
    for doc in documents:
        if doc["document_type"] in ["Insurance", "PUC"]:
            expiry = doc.get("expiry_date")
            if expiry:
                if isinstance(expiry, str):
                    expiry = datetime.fromisoformat(expiry)
                
                status_check = surepass_service.check_document_status(expiry)
                document_status[doc["document_type"].lower()] = status_check
                
                if status_check["needs_update"]:
                    needs_update = True
    
    return {
        "vehicle_id": vehicle_id,
        "registration_number": vehicle["registration_number"],
        "needs_update": needs_update,
        "document_status": document_status,
        "last_synced": vehicle.get("last_synced")
    }

@api_router.post("/vehicles/batch-check-documents")
async def batch_check_documents(
    current_user: dict = Depends(get_current_user)
):
    """
    Background job to check all vehicles for document expiry
    Returns vehicles that need update
    """
    vehicles = await db.vehicles.find({"is_deleted": False}).to_list(1000)
    
    needs_update = []
    
    for vehicle in vehicles:
        # Get current documents
        docs = await db.vehicle_documents.find({
            "vehicle_id": vehicle["id"],
            "is_current": True,
            "document_type": {"$in": ["Insurance", "PUC"]},
            "is_deleted": False
        }).to_list(5)
        
        vehicle_needs_update = False
        
        for doc in docs:
            if doc.get("expiry_date"):
                if isinstance(doc["expiry_date"], str):
                    expiry = datetime.fromisoformat(doc["expiry_date"])
                else:
                    expiry = doc["expiry_date"]
                
                days_left = (expiry - datetime.now(timezone.utc)).days
                
                if days_left <= 30:  # Needs update if <=30 days left
                    vehicle_needs_update = True
                    break
        
        if vehicle_needs_update:
            needs_update.append({
                "id": vehicle["id"],
                "registration_number": vehicle["registration_number"],
                "last_synced": vehicle.get("last_synced")
            })
    
    return {
        "total_vehicles": len(vehicles),
        "needs_update_count": len(needs_update),
        "vehicles_needing_update": needs_update
    }

@api_router.post("/surepass/fetch-challans")
async def fetch_challans_from_surepass(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Fetch challan details for a vehicle from Surepass API
    Updates existing challans (changes Unpaid to Paid if needed)
    Adds new challans if not present
    """
    registration_number = data.get("registration_number", "").upper().strip()
    
    if not registration_number:
        raise HTTPException(status_code=400, detail="Registration number is required")
    
    # Find the vehicle
    vehicle = await db.vehicles.find_one({
        "registration_number": {"$regex": f"^{registration_number}$", "$options": "i"},
        "is_deleted": False
    })
    
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    logger.info(f"Fetching challans for vehicle {registration_number}")
    
    # Call Surepass API
    result = await surepass_service.fetch_vehicle_challans(registration_number)
    
    if not result["success"]:
        raise HTTPException(status_code=400, detail=result.get("error", "Failed to fetch challan details"))
    
    challan_details = result.get("data", [])
    
    if not challan_details:
        return {
            "success": True,
            "vehicle_id": vehicle["id"],
            "registration_number": registration_number,
            "challans_found": 0,
            "challans_imported": 0,
            "challans_updated": 0,
            "challans": [],
            "message": "No challans found for this vehicle"
        }
    
    imported = 0
    updated = 0
    new_challans = []
    updated_challans = []
    
    for challan_data in challan_details:
        # Parse the challan data
        parsed_challan = surepass_service.parse_challan_data(challan_data, vehicle["id"])
        challan_number = parsed_challan["challan_number"]
        
        # Check if challan already exists
        existing = await db.challans.find_one({
            "challan_number": challan_number,
            "vehicle_id": vehicle["id"],
            "is_deleted": False
        })
        
        if existing:
            # 🔥 UPDATE EXISTING CHALLAN - Check if status changed from Unpaid to Paid
            existing_status = existing.get("status", "Unpaid")
            new_status = parsed_challan.get("status", "Unpaid")
            
            if existing_status == "Unpaid" and new_status == "Paid":
                # Update the challan status to Paid
                update_data = {
                    "status": new_status,
                    "payment_date": parsed_challan.get("payment_date") or datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user["user_id"]
                }
                
                await db.challans.update_one(
                    {"id": existing["id"]},
                    {"$set": update_data}
                )
                updated += 1
                updated_challans.append({
                    "challan_number": challan_number,
                    "old_status": existing_status,
                    "new_status": new_status,
                    "amount": parsed_challan.get("amount", 0)
                })
                logger.info(f"Updated challan {challan_number} status from {existing_status} to {new_status}")
            else:
                # Optionally update other fields if needed (like amount, etc.)
                # Some challans might have updated amounts
                needs_update = False
                update_fields = {}
                
                if existing.get("amount") != parsed_challan.get("amount"):
                    update_fields["amount"] = parsed_challan.get("amount", 0)
                    needs_update = True
                
                if existing.get("violation_type") != parsed_challan.get("violation_type"):
                    update_fields["violation_type"] = parsed_challan.get("violation_type")
                    needs_update = True
                
                if needs_update:
                    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
                    update_fields["updated_by"] = current_user["user_id"]
                    await db.challans.update_one(
                        {"id": existing["id"]},
                        {"$set": update_fields}
                    )
                    logger.info(f"Updated challan {challan_number} fields: {list(update_fields.keys())}")
        else:
            # Create new challan
            challan_id = str(uuid.uuid4())
            challan_dict = {
                "id": challan_id,
                **parsed_challan,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user["user_id"],
                "is_deleted": False
            }
            
            await db.challans.insert_one(challan_dict)
            imported += 1
            challan_dict.pop('_id', None)
            new_challans.append(challan_dict)
            logger.info(f"Imported new challan {challan_number}")
    
    # Log the API call
    log_entry = {
        "id": str(uuid.uuid4()),
        "registration_number": registration_number,
        "request_timestamp": datetime.now(timezone.utc).isoformat(),
        "response_data": result.get("raw_response", {}),
        "challans_found": len(challan_details),
        "challans_imported": imported,
        "challans_updated": updated,
        "is_successful": True,
        "created_by": current_user["user_id"]
    }
    await db.rc_verification_logs.insert_one(log_entry)
    
    # Build response message
    message_parts = []
    if imported > 0:
        message_parts.append(f"Imported {imported} new challans")
    if updated > 0:
        message_parts.append(f"Updated {updated} challans (Unpaid → Paid)")
    
    if not message_parts:
        message = "No new challans found and no status updates needed"
    else:
        message = ", ".join(message_parts)
    
    return {
        "success": True,
        "vehicle_id": vehicle["id"],
        "registration_number": registration_number,
        "challans_found": len(challan_details),
        "challans_imported": imported,
        "challans_updated": updated,
        "challans": new_challans,
        "updated_challans": updated_challans,
        "message": message
    }

# ==================== VEHICLE TRACKING ROUTES ====================

@api_router.get("/vehicle-tracking/live")
async def get_live_vehicle(imei: str, current_user: dict = Depends(get_current_user)):
    vehicle = await db.vehicle_live.find_one({"imei": imei}, {"_id": 0})

    if not vehicle:
        raise HTTPException(status_code=404, detail="No live data found")

    return vehicle

@api_router.get("/vehicle-tracking/history")
async def get_vehicle_history(
    imei: str,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    data = await db.vehicle_history.find(
        {"imei": imei},
        {"_id": 0}
    ).sort("timestamp", -1).limit(limit).to_list(limit)

    return {
        "count": len(data),
        "data": data[::-1]
    }

@api_router.get("/vehicle-tracking/timeline")
async def get_vehicle_timeline(
    imei: str,
    limit: int = 200,
    current_user: dict = Depends(get_current_user)
):
    records = await db.vehicle_history.find(
        {"imei": imei},
        {"_id": 0}
    ).sort("timestamp", 1).limit(limit).to_list(limit)

    if not records:
        return {"timeline": []}

    timeline = []
    last_state = None

    for i, r in enumerate(records):
        speed = r.get("speed", 0)
        if r.get("ignition") == "OFF":
            state = "ENGINE_OFF"
        elif speed == 0:
            state = "IDLE"
        else:
            state = "MOVING"

        current_time = datetime.fromisoformat(r["timestamp"])

        if i < len(records) - 1:
            next_time = datetime.fromisoformat(records[i + 1]["timestamp"])
            duration_min = (next_time - current_time).total_seconds() / 60
        else:
            duration_min = 0

        if state != last_state:
            timeline.append({
                "time": r["timestamp"],
                "type": state,
                "lat": r["lat"],
                "lng": r["lng"],
                "location": r.get("location"),  # ✅ LOCATION ADDED
                "duration_min": round(duration_min, 2)
            })
            last_state = state

    # 🔹 START
    timeline.insert(0, {
        "time": records[0]["timestamp"],
        "type": "START",
        "lat": records[0]["lat"],
        "lng": records[0]["lng"],
        "location": records[0].get("location")
    })

    # 🔹 END
    timeline.append({
        "time": records[-1]["timestamp"],
        "type": "END",
        "lat": records[-1]["lat"],
        "lng": records[-1]["lng"],
        "location": records[-1].get("location")
    })

    return {
        "imei": imei,
        "total_points": len(records),
        "timeline_points": len(timeline),
        "timeline": timeline
    }
    records = await db.vehicle_history.find(
        {"imei": imei},
        {"_id": 0}
    ).sort("timestamp", 1).limit(limit).to_list(limit)

    if not records:
        return {"timeline": []}

    timeline = []
    last_state = None

    for i, r in enumerate(records):
        speed = r.get("speed", 0)
        state = "STOP" if speed == 0 else "MOVING"

        current_time = datetime.fromisoformat(r["timestamp"])

        # 🔥 calculate duration
        if i < len(records) - 1:
            next_time = datetime.fromisoformat(records[i + 1]["timestamp"])
            duration_min = (next_time - current_time).total_seconds() / 60
        else:
            duration_min = 0

        if state != last_state:
            timeline.append({
                "time": r["timestamp"],
                "type": state,
                "lat": r["lat"],
                "lng": r["lng"],
                "duration_min": round(duration_min, 2)  # 🔥 NEW
            })
            last_state = state

    # 🔹 START
    timeline.insert(0, {
        "time": records[0]["timestamp"],
        "type": "START"
    })

    # 🔹 END
    timeline.append({
        "time": records[-1]["timestamp"],
        "type": "END"
    })

    return {
        "imei": imei,
        "total_points": len(records),
        "timeline_points": len(timeline),
        "timeline": timeline
    }
# ==================== VEHICLE DOCUMENT ROUTES (VERSIONED) ====================

@api_router.get("/vehicle-documents")
async def get_vehicle_documents(
    vehicle_id: Optional[str] = None,
    current_only: bool = True,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    """
    Get vehicle documents with optional filters
    """
    query = {"is_deleted": False}
    
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    # Only filter by is_current if current_only is True
    if current_only:
        query["is_current"] = True
    
    logger.info(f"Vehicle documents query: {query}")
    
    docs = await db.vehicle_documents.find(
        query, 
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    logger.info(f"Found {len(docs)} documents")
    
    # Ensure dates are properly formatted for frontend
    for doc in docs:
        if doc.get("expiry_date"):
            # Keep as ISO string, frontend will parse
            pass
    
    return {"data": docs, "total": len(docs)}

@api_router.get("/vehicle-documents/{document_id}/history")
async def get_document_history(document_id: str, current_user: dict = Depends(get_current_user)):
    # Get current document
    doc = await db.vehicle_documents.find_one({"id": document_id, "is_deleted": False}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Build version history
    history = [doc]
    current_doc = doc
    
    # Traverse backwards through versions
    while current_doc.get("previous_version_id"):
        prev_doc = await db.vehicle_documents.find_one({"id": current_doc["previous_version_id"]}, {"_id": 0})
        if prev_doc:
            history.append(prev_doc)
            current_doc = prev_doc
        else:
            break
    
    return {"vehicle_id": doc["vehicle_id"], "document_type": doc["document_type"], "history": history}
   
@api_router.post("/upload-document")
async def upload_document(file: UploadFile = File(...)):
    try:
        # Get file extension
        file_extension = file.filename.split('.')[-1].lower() if '.' in file.filename else 'pdf'
        
        # Determine resource type based on file extension
        if file_extension in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
            resource_type = "image"
        else:
            resource_type = "raw"
        
        # Upload to Cloudinary
        result = cloudinary.uploader.upload(
            file.file,
            folder="vehicle-documents",
            resource_type=resource_type,
            transformation=[
                {"quality": "auto", "fetch_format": "auto", "width": 1200, "crop": "limit"}
            ] if resource_type == "image" else None
        )
        
        return {
            "url": result["secure_url"],
            "public_id": result["public_id"],
            "file_type": file_extension  # Return the file type
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
# Add these endpoints after your existing document endpoints

@api_router.post("/vehicles/{vehicle_id}/upload-rc")
async def upload_rc_document(
    vehicle_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload RC document for a vehicle"""
    try:
        # Find the vehicle
        vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        # Get file extension from filename
        file_extension = file.filename.split('.')[-1].lower() if '.' in file.filename else ''
        
        # Determine resource type based on file extension
        if file_extension in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
            resource_type = "image"
        else:
            resource_type = "raw"
        
        # Upload to Cloudinary
        result = cloudinary.uploader.upload(
            file.file,
            folder=f"vehicles/{vehicle_id}/rc",
            resource_type=resource_type,
            transformation=[
                {"quality": "auto", "fetch_format": "auto", "width": 1200, "crop": "limit"}
            ] if resource_type == "image" else None
        )
        
        # Update vehicle with RC document URL and file type
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$set": {
                "rc_document_url": result["secure_url"],
                "rc_document_public_id": result["public_id"],
                "rc_document_file_type": file_extension if file_extension else 'bin',
                "rc_document_uploaded_at": datetime.now(timezone.utc).isoformat(),
                "rc_document_uploaded_by": current_user["user_id"]
            }}
        )
        
        return {
            "success": True,
            "url": result["secure_url"],
            "public_id": result["public_id"],
            "file_type": file_extension,
            "message": "RC document uploaded successfully"
        }
        
    except Exception as e:
        logger.error(f"RC upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/vehicles/{vehicle_id}/delete-rc")
async def delete_rc_document(
    vehicle_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete RC document for a vehicle"""
    try:
        vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        public_id = vehicle.get("rc_document_public_id")
        if public_id:
            # Delete from Cloudinary
            cloudinary.uploader.destroy(public_id)
        
        # Remove RC document fields from vehicle
        await db.vehicles.update_one(
            {"id": vehicle_id},
            {"$unset": {
                "rc_document_url": "",
                "rc_document_public_id": "",
                "rc_document_uploaded_at": "",
                "rc_document_uploaded_by": ""
            }}
        )
        
        return {"success": True, "message": "RC document deleted successfully"}
        
    except Exception as e:
        logger.error(f"RC delete error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/vehicles/{vehicle_id}/upload-document/{document_id}")
async def upload_document_file(
    vehicle_id: str,
    document_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload file for a specific document (Insurance/PUC)"""
    try:
        # Find the document
        document = await db.vehicle_documents.find_one({
            "id": document_id,
            "vehicle_id": vehicle_id,
            "is_deleted": False
        })
        
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        
        # Get file extension from filename
        file_extension = file.filename.split('.')[-1].lower() if '.' in file.filename else ''
        
        # Determine resource type based on file extension
        if file_extension in ['jpg', 'jpeg', 'png', 'gif', 'webp']:
            resource_type = "image"
        else:
            resource_type = "raw"
        
        # Upload to Cloudinary
        folder = f"vehicles/{vehicle_id}/{document['document_type'].lower()}"
        
        result = cloudinary.uploader.upload(
            file.file,
            folder=folder,
            resource_type=resource_type,
            transformation=[
                {"quality": "auto", "fetch_format": "auto", "width": 1200, "crop": "limit"}
            ] if resource_type == "image" else None
        )
        
        # Update document with file URL and file type
        await db.vehicle_documents.update_one(
            {"id": document_id},
            {"$set": {
                "file_url": result["secure_url"],
                "file_public_id": result["public_id"],
                "file_type": file_extension,
                "file_uploaded_at": datetime.now(timezone.utc).isoformat(),
                "file_uploaded_by": current_user["user_id"]
            }}
        )
        
        return {
            "success": True,
            "url": result["secure_url"],
            "public_id": result["public_id"],
            "file_type": file_extension,
            "message": f"{document['document_type']} document uploaded successfully"
        }
        
    except Exception as e:
        logger.error(f"Document upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.delete("/vehicles/{vehicle_id}/delete-document/{document_id}")
async def delete_document_file(
    vehicle_id: str,
    document_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete file for a specific document"""
    try:
        document = await db.vehicle_documents.find_one({
            "id": document_id,
            "vehicle_id": vehicle_id,
            "is_deleted": False
        })
        
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        
        public_id = document.get("file_public_id")
        if public_id:
            # Delete from Cloudinary
            cloudinary.uploader.destroy(public_id)
        
        # Remove file fields from document
        await db.vehicle_documents.update_one(
            {"id": document_id},
            {"$unset": {
                "file_url": "",
                "file_public_id": "",
                "file_uploaded_at": "",
                "file_uploaded_by": ""
            }}
        )
        
        return {"success": True, "message": "Document file deleted successfully"}
        
    except Exception as e:
        logger.error(f"Document delete error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/vehicles/{vehicle_id}/download-document/{document_id}")
async def download_document_file(
    vehicle_id: str,
    document_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download document file directly"""
    try:
        # Find the document
        document = await db.vehicle_documents.find_one({
            "id": document_id,
            "vehicle_id": vehicle_id,
            "is_deleted": False
        })
        
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        
        file_url = document.get("file_url")
        if not file_url:
            raise HTTPException(status_code=404, detail="No file uploaded for this document")
        
        # Get vehicle registration number
        vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "registration_number": 1})
        registration = vehicle.get("registration_number", "unknown").replace("/", "_").replace(" ", "_").replace("-", "_")
        
        doc_type = document.get("document_type", "document").lower()
        
        logger.info(f"Downloading {doc_type} document from: {file_url}")
        
        # Fetch the file from Cloudinary
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(file_url)
            
            if response.status_code != 200:
                logger.error(f"Failed to fetch file from Cloudinary: {response.status_code}")
                raise HTTPException(status_code=404, detail="File not found on Cloudinary")
            
            # Get the stored file type
            file_type = document.get("file_type", "")
            
            # If stored file type exists, use it
            if file_type:
                if file_type in ['pdf']:
                    content_type = 'application/pdf'
                    filename = f"{registration}_{doc_type}.pdf"
                elif file_type in ['jpg', 'jpeg']:
                    content_type = 'image/jpeg'
                    filename = f"{registration}_{doc_type}.jpg"
                elif file_type in ['png']:
                    content_type = 'image/png'
                    filename = f"{registration}_{doc_type}.png"
                elif file_type in ['gif']:
                    content_type = 'image/gif'
                    filename = f"{registration}_{doc_type}.gif"
                else:
                    content_type = 'application/octet-stream'
                    filename = f"{registration}_{doc_type}.{file_type}"
            else:
                # Fallback: determine from content type
                content_type = response.headers.get('content-type', 'application/octet-stream')
                
                if 'pdf' in content_type:
                    filename = f"{registration}_{doc_type}.pdf"
                    content_type = 'application/pdf'
                elif 'jpeg' in content_type or 'jpg' in content_type:
                    filename = f"{registration}_{doc_type}.jpg"
                    content_type = 'image/jpeg'
                elif 'png' in content_type:
                    filename = f"{registration}_{doc_type}.png"
                    content_type = 'image/png'
                elif 'gif' in content_type:
                    filename = f"{registration}_{doc_type}.gif"
                    content_type = 'image/gif'
                else:
                    # Try to get extension from URL
                    url_parts = file_url.split('.')
                    ext = url_parts[-1].split('?')[0] if len(url_parts) > 1 else 'bin'
                    filename = f"{registration}_{doc_type}.{ext}"
            
            logger.info(f"Serving file with content-type: {content_type}, filename: {filename}")
            
            # Return the file directly
            return StreamingResponse(
                iter([response.content]),
                media_type=content_type,
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "Content-Length": str(len(response.content)),
                    "Cache-Control": "no-cache",
                    "Access-Control-Expose-Headers": "Content-Disposition"
                }
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Document download error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to download file: {str(e)}")

# ==================== SOLD VEHICLE AGREEMENT ROUTES ====================

@api_router.post("/vehicles/{vehicle_id}/sold-agreement")
async def create_sold_agreement(
    vehicle_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Create or update a sold vehicle agreement with optional document upload
    """
    try:
        # Find the vehicle
        vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        # Check if vehicle is marked as sold
        if not vehicle.get("sold", False):
            raise HTTPException(status_code=400, detail="Vehicle is not marked as sold")
        
        # Check if agreement already exists
        existing = await db.sold_vehicle_agreements.find_one(
            {"vehicle_id": vehicle_id, "is_deleted": False}
        )
        
        if existing:
            # Update existing agreement
            update_data = {
                "buyer_name": data.get("buyer_name"),
                "buyer_phone": data.get("buyer_phone"),
                "agreement_date": datetime.fromisoformat(data.get("agreement_date", datetime.now(timezone.utc).isoformat())),
                "notes": data.get("notes"),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user["user_id"]
            }
            
            await db.sold_vehicle_agreements.update_one(
                {"id": existing["id"]},
                {"$set": update_data}
            )
            
            return {
                "success": True,
                "agreement_id": existing["id"],
                "message": "Sold vehicle agreement updated successfully"
            }
        else:
            # Create new agreement
            agreement_id = str(uuid.uuid4())
            agreement = {
                "id": agreement_id,
                "vehicle_id": vehicle_id,
                "vehicle_registration": vehicle["registration_number"],
                "buyer_name": data.get("buyer_name"),
                "buyer_phone": data.get("buyer_phone"),
                "agreement_date": datetime.fromisoformat(data.get("agreement_date", datetime.now(timezone.utc).isoformat())),
                "agreement_url": data.get("agreement_url"),
                "agreement_public_id": data.get("agreement_public_id"),
                "agreement_file_type": data.get("agreement_file_type"),
                "notes": data.get("notes"),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user["user_id"],
                "is_deleted": False
            }
            
            await db.sold_vehicle_agreements.insert_one(agreement)
            
            # Update vehicle with agreement reference
            await db.vehicles.update_one(
                {"id": vehicle_id},
                {"$set": {
                    "sold_agreement_id": agreement_id,
                    "sold_date": datetime.now(timezone.utc).isoformat(),
                    "sold_agreement_created": True
                }}
            )
            
            return {
                "success": True,
                "agreement_id": agreement_id,
                "message": "Sold vehicle agreement created successfully"
            }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating sold agreement: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to create agreement: {str(e)}")
     
@api_router.post("/vehicles/{vehicle_id}/upload-agreement")
async def upload_agreement_document(
    vehicle_id: str, 
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload agreement document for sold vehicle"""
    try:
        # Validate file type
        if file.content_type != "application/pdf":
            raise HTTPException(status_code=400, detail="Only PDF files are allowed")
        
        # Validate file size (10MB max)
        file_content = await file.read()
        if len(file_content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File size must be less than 10MB")
        
        # Reset file pointer
        await file.seek(0)
        
        # Upload to Cloudinary
        upload_result = cloudinary.uploader.upload(
            file.file,
            resource_type="raw",
            folder=f"sold_agreements/{vehicle_id}"
        )
        
        agreement_url = upload_result.get("secure_url")
        public_id = upload_result.get("public_id")
        
        # Find existing agreement
        existing = await db.sold_vehicle_agreements.find_one(
            {"vehicle_id": vehicle_id, "is_deleted": False}
        )
        
        if existing:
            # Update existing agreement with document URL
            await db.sold_vehicle_agreements.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "agreement_url": agreement_url,
                    "agreement_public_id": public_id,
                    "agreement_file_type": "pdf",
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user["user_id"]
                }}
            )
        else:
            # Create agreement with document
            agreement_id = str(uuid.uuid4())
            vehicle = await db.vehicles.find_one({"id": vehicle_id})
            
            agreement = {
                "id": agreement_id,
                "vehicle_id": vehicle_id,
                "vehicle_registration": vehicle["registration_number"] if vehicle else "Unknown",
                "agreement_url": agreement_url,
                "agreement_public_id": public_id,
                "agreement_file_type": "pdf",
                "agreement_date": datetime.now(timezone.utc).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": current_user["user_id"],
                "is_deleted": False
            }
            
            await db.sold_vehicle_agreements.insert_one(agreement)
        
        return {
            "success": True,
            "url": agreement_url,
            "message": "Agreement document uploaded successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading agreement: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to upload document: {str(e)}")
    
@api_router.get("/vehicles/{vehicle_id}/sold-agreement")
async def get_sold_agreement(vehicle_id: str, current_user: dict = Depends(get_current_user)):
    """Get sold agreement details for a vehicle"""
    try:
        agreement = await db.sold_vehicle_agreements.find_one(
            {"vehicle_id": vehicle_id, "is_deleted": False}, 
            {"_id": 0}
        )
        
        if not agreement:
            return {"has_agreement": False}
        
        # Also get vehicle sold status
        vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "sold": 1})
        
        return {
            "has_agreement": True,
            "agreement": agreement,
            "vehicle_sold": vehicle.get("sold", False) if vehicle else False
        }
        
    except Exception as e:
        logger.error(f"Error fetching agreement: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
@api_router.get("/vehicles/{vehicle_id}/download-agreement")
async def download_sold_agreement(
    vehicle_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download sold agreement PDF"""
    try:
        agreement = await db.sold_vehicle_agreements.find_one(
            {"vehicle_id": vehicle_id, "is_deleted": False}
        )
        
        if not agreement or not agreement.get("agreement_url"):
            raise HTTPException(status_code=404, detail="Agreement not found")
        
        agreement_url = agreement["agreement_url"]
        
        # Fetch the file from Cloudinary
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(agreement_url)
            
            if response.status_code != 200:
                logger.error(f"Failed to fetch agreement from Cloudinary: {response.status_code}")
                raise HTTPException(status_code=404, detail="Agreement file not found")
            
            vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "registration_number": 1})
            registration = vehicle.get("registration_number", "unknown").replace("/", "_").replace(" ", "_").replace("-", "_")
            filename = f"{registration}_sold_agreement.pdf"
            
            return StreamingResponse(
                iter([response.content]),
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "Content-Length": str(len(response.content)),
                    "Cache-Control": "no-cache"
                }
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading agreement: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to download: {str(e)}")
    
@api_router.delete("/vehicles/{vehicle_id}/sold-agreement")
async def delete_sold_agreement(
    vehicle_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete sold vehicle agreement"""
    try:
        agreement = await db.sold_vehicle_agreements.find_one(
            {"vehicle_id": vehicle_id, "is_deleted": False}
        )
        
        if not agreement:
            raise HTTPException(status_code=404, detail="No agreement found")
        
        # Delete file from Cloudinary if exists
        if agreement.get("agreement_public_id"):
            try:
                cloudinary.uploader.destroy(agreement["agreement_public_id"], resource_type="raw")
            except Exception as e:
                logger.warning(f"Failed to delete from Cloudinary: {e}")
        
        # Soft delete the agreement
        await db.sold_vehicle_agreements.update_one(
            {"id": agreement["id"]},
            {"$set": {
                "is_deleted": True, 
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user["user_id"]
            }}
        )
        
        return {"success": True, "message": "Agreement deleted successfully"}
        
    except Exception as e:
        logger.error(f"Agreement delete error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.patch("/vehicles/{vehicle_id}/mark-unsold")
async def mark_vehicle_unsold(
    vehicle_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Mark vehicle as unsold and delete all associated sold data"""
    try:
        # Update vehicle sold status
        result = await db.vehicles.update_one(
            {"id": vehicle_id, "is_deleted": False},
            {"$set": {
                "sold": False,
                "sold_date": None,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user["user_id"]
            }}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        # Find and delete agreement
        agreement = await db.sold_vehicle_agreements.find_one(
            {"vehicle_id": vehicle_id, "is_deleted": False}
        )
        
        if agreement:
            # Delete file from Cloudinary if exists
            if agreement.get("agreement_public_id"):
                try:
                    cloudinary.uploader.destroy(agreement["agreement_public_id"], resource_type="raw")
                except Exception as e:
                    logger.warning(f"Failed to delete from Cloudinary: {e}")
            
            # Delete agreement
            await db.sold_vehicle_agreements.update_one(
                {"id": agreement["id"]},
                {"$set": {
                    "is_deleted": True,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user["user_id"]
                }}
            )
        
        return {"success": True, "message": "Vehicle marked as unsold successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error marking as unsold: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.patch("/vehicles/{vehicle_id}/sold-status")
async def update_vehicle_sold_status(
    vehicle_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update the sold status of a vehicle"""
    sold_status = data.get("sold", False)
    
    result = await db.vehicles.update_one(
        {"id": vehicle_id, "is_deleted": False},
        {"$set": {
            "sold": sold_status,
            "sold_date": datetime.now(timezone.utc).isoformat() if sold_status else None,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user["user_id"]
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    return {
        "success": True,
        "message": f"Vehicle sold status updated to {sold_status}",
        "sold": sold_status
    }

@api_router.get("/vehicles/{vehicle_id}/download-rc")
async def download_rc_document(
    vehicle_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Download RC document file directly"""
    try:
        # Find the vehicle
        vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        rc_url = vehicle.get("rc_document_url")
        if not rc_url:
            raise HTTPException(status_code=404, detail="No RC document uploaded")
        
        logger.info(f"Downloading RC document from: {rc_url}")
        
        # Fetch the file from Cloudinary
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(rc_url)
            
            if response.status_code != 200:
                logger.error(f"Failed to fetch file from Cloudinary: {response.status_code}")
                raise HTTPException(status_code=404, detail="File not found on Cloudinary")
            
            # Get the stored file type or determine from content type
            file_type = vehicle.get("rc_document_file_type", "")
            
            # If stored file type exists, use it
            if file_type:
                if file_type in ['pdf']:
                    content_type = 'application/pdf'
                    filename = f"{vehicle['registration_number'].replace('/', '_').replace(' ', '_').replace('-', '_')}_RC.pdf"
                elif file_type in ['jpg', 'jpeg']:
                    content_type = 'image/jpeg'
                    filename = f"{vehicle['registration_number'].replace('/', '_').replace(' ', '_').replace('-', '_')}_RC.jpg"
                elif file_type in ['png']:
                    content_type = 'image/png'
                    filename = f"{vehicle['registration_number'].replace('/', '_').replace(' ', '_').replace('-', '_')}_RC.png"
                elif file_type in ['gif']:
                    content_type = 'image/gif'
                    filename = f"{vehicle['registration_number'].replace('/', '_').replace(' ', '_').replace('-', '_')}_RC.gif"
                else:
                    content_type = 'application/octet-stream'
                    filename = f"{vehicle['registration_number'].replace('/', '_').replace(' ', '_').replace('-', '_')}_RC.{file_type}"
            else:
                # Fallback: determine from content type
                content_type = response.headers.get('content-type', 'application/octet-stream')
                registration = vehicle["registration_number"].replace("/", "_").replace(" ", "_").replace("-", "_")
                
                if 'pdf' in content_type:
                    filename = f"{registration}_RC.pdf"
                    content_type = 'application/pdf'
                elif 'jpeg' in content_type or 'jpg' in content_type:
                    filename = f"{registration}_RC.jpg"
                    content_type = 'image/jpeg'
                elif 'png' in content_type:
                    filename = f"{registration}_RC.png"
                    content_type = 'image/png'
                elif 'gif' in content_type:
                    filename = f"{registration}_RC.gif"
                    content_type = 'image/gif'
                else:
                    # Try to get extension from URL
                    url_parts = rc_url.split('.')
                    ext = url_parts[-1].split('?')[0] if len(url_parts) > 1 else 'bin'
                    filename = f"{registration}_RC.{ext}"
            
            logger.info(f"Serving file with content-type: {content_type}, filename: {filename}")
            
            # Return the file directly
            return StreamingResponse(
                iter([response.content]),
                media_type=content_type,
                headers={
                    "Content-Disposition": f"attachment; filename={filename}",
                    "Content-Length": str(len(response.content)),
                    "Cache-Control": "no-cache",
                    "Access-Control-Expose-Headers": "Content-Disposition"
                }
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"RC download error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to download file: {str(e)}")

@api_router.patch("/vehicles/{vehicle_id}/sold-status")
async def update_vehicle_sold_status(
    vehicle_id: str,
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update the sold status of a vehicle"""
    sold_status = data.get("sold", False)
    
    result = await db.vehicles.update_one(
        {"id": vehicle_id, "is_deleted": False},
        {"$set": {
            "sold": sold_status,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user["user_id"]
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    return {
        "success": True,
        "message": f"Vehicle sold status updated to {sold_status}",
        "sold": sold_status
    }
# Add this PUT endpoint after your GET /vehicle-documents endpoint and before the history endpoint

@api_router.put("/vehicle-documents/{document_id}")
async def update_vehicle_document(
    document_id: str,
    document_data: VehicleDocumentCreate,
    current_user: dict = Depends(get_current_user)
):
    """
    Update an existing vehicle document (direct update, no versioning)
    """
    # Check if document exists
    existing_doc = await db.vehicle_documents.find_one({"id": document_id, "is_deleted": False})
    if not existing_doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Prepare update data
    update_data = document_data.model_dump()
    
    # Convert datetime fields to ISO format
    if isinstance(update_data["issue_date"], datetime):
        update_data["issue_date"] = update_data["issue_date"].isoformat()
    if isinstance(update_data["expiry_date"], datetime):
        update_data["expiry_date"] = update_data["expiry_date"].isoformat()
    
    # Handle premium conversion
    if update_data.get("premium"):
        update_data["premium"] = float(update_data["premium"])
    
    # Add update metadata
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["updated_by"] = current_user["user_id"]
    
    # Remove fields that shouldn't be updated
    update_data.pop("created_at", None)
    update_data.pop("created_by", None)
    update_data.pop("version", None)
    update_data.pop("is_current", None)
    update_data.pop("previous_version_id", None)
    
    # Update the document
    result = await db.vehicle_documents.update_one(
        {"id": document_id, "is_deleted": False},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="No changes made to document")
    
    # 🔥 NEW: Update vehicle's main fields if this is a current document
    if existing_doc.get("is_current") and not existing_doc.get("is_past_document"):
        vehicle_update = {}
        doc_type = existing_doc.get("document_type")
        expiry_date = update_data.get("expiry_date", existing_doc.get("expiry_date"))
        
        if doc_type == "Insurance":
            vehicle_update["insurance_expiry"] = expiry_date
            vehicle_update["insurance_company"] = update_data.get("provider", existing_doc.get("provider"))
            vehicle_update["insurance_policy_number"] = update_data.get("policy_number", existing_doc.get("policy_number"))
        elif doc_type == "PUC":
            vehicle_update["puc_expiry"] = expiry_date
            vehicle_update["pucc_number"] = update_data.get("policy_number", existing_doc.get("policy_number"))
        elif doc_type == "Fitness":
            vehicle_update["fit_up_to"] = expiry_date
        elif doc_type == "Tax":
            vehicle_update["tax_upto"] = expiry_date
        
        if vehicle_update:
            vehicle_update["updated_at"] = datetime.now(timezone.utc).isoformat()
            vehicle_update["updated_by"] = current_user["user_id"]
            await db.vehicles.update_one(
                {"id": existing_doc["vehicle_id"]},
                {"$set": vehicle_update}
            )
            logger.info(f"Updated vehicle {existing_doc['vehicle_id']} {doc_type} expiry to {expiry_date}")
    
    return {
        "success": True,
        "message": "Document updated successfully",
        "document_id": document_id
    }

@api_router.delete("/vehicle-documents/{document_id}/version")
async def delete_document_version(
    document_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Delete a specific version of a document
    """
    # Check if document exists
    existing_doc = await db.vehicle_documents.find_one({"id": document_id, "is_deleted": False})
    if not existing_doc:
        raise HTTPException(status_code=404, detail="Document version not found")
    
    # If this is the current version, we need to handle carefully
    if existing_doc.get("is_current"):
        # Check if there's a previous version
        previous_version_id = existing_doc.get("previous_version_id")
        if previous_version_id:
            # Mark previous version as current
            await db.vehicle_documents.update_one(
                {"id": previous_version_id},
                {"$set": {"is_current": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
    
    # Soft delete this version
    result = await db.vehicle_documents.update_one(
        {"id": document_id},
        {
            "$set": {
                "is_deleted": True,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user["user_id"]
            }
        }
    )
    
    return {
        "success": True,
        "message": "Document version deleted successfully"
    }    

@api_router.delete("/vehicle-documents/{document_id}")
async def delete_vehicle_document(
    document_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Delete a vehicle document (soft delete all versions)
    """
    try:
        # Find the document
        document = await db.vehicle_documents.find_one({"id": document_id, "is_deleted": False})
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        
        # Soft delete the document
        await db.vehicle_documents.update_one(
            {"id": document_id},
            {
                "$set": {
                    "is_deleted": True,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user["user_id"]
                }
            }
        )
        
        return {
            "success": True,
            "message": "Document deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting vehicle document: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to delete document: {str(e)}")\

@api_router.get("/export/vehicle-documents/excel")
async def export_vehicle_documents_excel(
    vehicle_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Export vehicle documents to a professional Excel file
    """
    try:
        # Build query
        query = {"is_deleted": False}
        if vehicle_id:
            query["vehicle_id"] = vehicle_id
        
        # Get all documents
        documents = await db.vehicle_documents.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
        
        if not documents:
            raise HTTPException(status_code=404, detail="No documents found to export")
        
        # Get vehicle information for reference
        vehicle_ids = list(set(doc["vehicle_id"] for doc in documents if doc.get("vehicle_id")))
        vehicles = []
        if vehicle_ids:
            vehicles = await db.vehicles.find(
                {"id": {"$in": vehicle_ids}},
                {"_id": 0, "id": 1, "registration_number": 1, "brand": 1, "model": 1}
            ).to_list(100)
        
        vehicle_map = {v["id"]: v for v in vehicles}
        
        # Prepare data for Excel
        excel_data = []
        for doc in documents:
            vehicle = vehicle_map.get(doc.get("vehicle_id"), {})
            
            # Calculate days until expiry
            days_left = None
            expiry_status = "Unknown"
            if doc.get("expiry_date"):
                try:
                    expiry = datetime.fromisoformat(doc["expiry_date"]) if isinstance(doc["expiry_date"], str) else doc["expiry_date"]
                    days_left = (expiry - datetime.now(timezone.utc)).days
                    
                    if days_left <= 0:
                        expiry_status = "Expired"
                    elif days_left <= 3:
                        expiry_status = "Critical"
                    elif days_left <= 7:
                        expiry_status = "Urgent"
                    elif days_left <= 15:
                        expiry_status = "Warning"
                    elif days_left <= 30:
                        expiry_status = "Soon"
                    else:
                        expiry_status = "Safe"
                except Exception as e:
                    logger.error(f"Error calculating days left: {e}")
            
            # Format dates safely
            issue_date = ""
            if doc.get("issue_date"):
                try:
                    if isinstance(doc["issue_date"], str):
                        issue_date = datetime.fromisoformat(doc["issue_date"].replace('Z', '+00:00')).strftime('%Y-%m-%d')
                    else:
                        issue_date = doc["issue_date"].strftime('%Y-%m-%d')
                except:
                    issue_date = str(doc["issue_date"])
            
            expiry_date = ""
            if doc.get("expiry_date"):
                try:
                    if isinstance(doc["expiry_date"], str):
                        expiry_date = datetime.fromisoformat(doc["expiry_date"].replace('Z', '+00:00')).strftime('%Y-%m-%d')
                    else:
                        expiry_date = doc["expiry_date"].strftime('%Y-%m-%d')
                except:
                    expiry_date = str(doc["expiry_date"])
            
            created_at = ""
            if doc.get("created_at"):
                try:
                    if isinstance(doc["created_at"], str):
                        created_at = datetime.fromisoformat(doc["created_at"].replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        created_at = doc["created_at"].strftime('%Y-%m-%d %H:%M:%S')
                except:
                    created_at = str(doc["created_at"])
            
            updated_at = ""
            if doc.get("updated_at"):
                try:
                    if isinstance(doc["updated_at"], str):
                        updated_at = datetime.fromisoformat(doc["updated_at"].replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        updated_at = doc["updated_at"].strftime('%Y-%m-%d %H:%M:%S')
                except:
                    updated_at = str(doc["updated_at"])
            
            excel_data.append({
                "Registration Number": vehicle.get("registration_number", "Unknown"),
                "Vehicle Model": f"{vehicle.get('brand', '')} {vehicle.get('model', '')}".strip(),
                "Document Type": doc.get("document_type", ""),
                "Custom Name": doc.get("custom_document_name", ""),
                "Policy Number": doc.get("policy_number", ""),
                "Provider": doc.get("provider", ""),
                "Phone Number": doc.get("phone_number", ""),
                "Issue Date": issue_date,
                "Expiry Date": expiry_date,
                "Days Left": str(days_left) if days_left is not None else "N/A",
                "Status": doc.get("status", ""),
                "Expiry Status": expiry_status,
                "Version": str(doc.get("version", 1)),
                "Is Current": "Yes" if doc.get("is_current", False) else "No",
                "Premium/Fee (Rs)": str(doc.get("premium", "")) if doc.get("premium") else "",
                "Coverage": doc.get("coverage", ""),
                "Created At": created_at,
                "Updated At": updated_at,
            })
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Create Excel file in memory using openpyxl
        output = io.BytesIO()
        
        # Use openpyxl engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Vehicle Documents', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Vehicle Documents']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet
            summary_data = {
                'Metric': [
                    'Total Documents',
                    'Active Documents',
                    'Expired Documents',
                    'Critical (≤3 days)',
                    'Urgent (≤7 days)',
                    'Warning (≤15 days)',
                    'Soon (≤30 days)',
                    'Safe (>30 days)',
                    'Vehicles Covered',
                    'Export Date'
                ],
                'Value': [
                    str(len(documents)),
                    str(len([d for d in documents if d.get('status') == 'Active'])),
                    str(len([d for d in documents if d.get('status') == 'Expired'])),
                    str(len([d for d in excel_data if d.get('Expiry Status') == 'Critical'])),
                    str(len([d for d in excel_data if d.get('Expiry Status') == 'Urgent'])),
                    str(len([d for d in excel_data if d.get('Expiry Status') == 'Warning'])),
                    str(len([d for d in excel_data if d.get('Expiry Status') == 'Soon'])),
                    str(len([d for d in excel_data if d.get('Expiry Status') == 'Safe'])),
                    str(len(vehicle_ids)),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Auto-adjust summary sheet columns
            summary_sheet = writer.sheets['Summary']
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 20
        
        output.seek(0)
        
        # Generate filename
        if vehicle_id:
            vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "registration_number": 1})
            vehicle_reg = vehicle["registration_number"].replace(" ", "_") if vehicle else "vehicle"
            filename = f"{vehicle_reg}_documents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"all_vehicles_documents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting vehicle documents: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export documents: {str(e)}")
        
@api_router.post("/export/vehicle-documents/current-view/excel")
async def export_current_view_excel(
    filter_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Export the current filtered view of documents to Excel
    """
    try:
        documents = filter_data.get("documents", [])
        vehicle_map = filter_data.get("vehicle_map", {})
        
        if not documents:
            raise HTTPException(status_code=404, detail="No documents to export")
        
        # Prepare data for Excel
        excel_data = []
        for doc in documents:
            vehicle = vehicle_map.get(doc.get("vehicle_id"), {})
            
            # Calculate days until expiry
            days_left = None
            expiry_status = "Unknown"
            if doc.get("expiry_date"):
                try:
                    expiry = datetime.fromisoformat(doc["expiry_date"]) if isinstance(doc["expiry_date"], str) else doc["expiry_date"]
                    days_left = (expiry - datetime.now(timezone.utc)).days
                    
                    if days_left <= 0:
                        expiry_status = "Expired"
                    elif days_left <= 3:
                        expiry_status = "Critical"
                    elif days_left <= 7:
                        expiry_status = "Urgent"
                    elif days_left <= 15:
                        expiry_status = "Warning"
                    elif days_left <= 30:
                        expiry_status = "Soon"
                    else:
                        expiry_status = "Safe"
                except:
                    pass
            
            # Format dates safely
            issue_date = ""
            if doc.get("issue_date"):
                try:
                    if isinstance(doc["issue_date"], str):
                        issue_date = datetime.fromisoformat(doc["issue_date"].replace('Z', '+00:00')).strftime('%Y-%m-%d')
                    else:
                        issue_date = doc["issue_date"].strftime('%Y-%m-%d')
                except:
                    issue_date = str(doc["issue_date"])
            
            expiry_date = ""
            if doc.get("expiry_date"):
                try:
                    if isinstance(doc["expiry_date"], str):
                        expiry_date = datetime.fromisoformat(doc["expiry_date"].replace('Z', '+00:00')).strftime('%Y-%m-%d')
                    else:
                        expiry_date = doc["expiry_date"].strftime('%Y-%m-%d')
                except:
                    expiry_date = str(doc["expiry_date"])
            
            created_at = ""
            if doc.get("created_at"):
                try:
                    if isinstance(doc["created_at"], str):
                        created_at = datetime.fromisoformat(doc["created_at"].replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        created_at = doc["created_at"].strftime('%Y-%m-%d %H:%M:%S')
                except:
                    created_at = str(doc["created_at"])
            
            excel_data.append({
                "Registration Number": vehicle.get("registration_number", "Unknown"),
                "Vehicle Model": f"{vehicle.get('brand', '')} {vehicle.get('model', '')}".strip(),
                "Document Type": doc.get("document_type", ""),
                "Custom Name": doc.get("custom_document_name", ""),
                "Policy Number": doc.get("policy_number", ""),
                "Provider": doc.get("provider", ""),
                "Phone Number": doc.get("phone_number", ""),
                "Issue Date": issue_date,
                "Expiry Date": expiry_date,
                "Days Left": str(days_left) if days_left is not None else "N/A",
                "Status": doc.get("status", ""),
                "Expiry Status": expiry_status,
                "Version": str(doc.get("version", 1)),
                "Is Current": "Yes" if doc.get("is_current", False) else "No",
                "Premium/Fee (Rs)": str(doc.get("premium", "")) if doc.get("premium") else "",
                "Coverage": doc.get("coverage", ""),
                "Created At": created_at,
            })
        
        # Create DataFrame and Excel file
        df = pd.DataFrame(excel_data)
        output = io.BytesIO()
        
        # Use openpyxl engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Vehicle Documents', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Vehicle Documents']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=documents_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting current view: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export documents: {str(e)}")
       
@api_router.post("/challans")
async def create_challan(challan_data: ChallanCreate, current_user: dict = Depends(get_current_user)):
    challan_id = str(uuid.uuid4())
    challan = Challan(id=challan_id, **challan_data.model_dump(), created_by=current_user["user_id"])
    
    challan_dict = challan.model_dump()
    challan_dict["created_at"] = challan_dict["created_at"].isoformat()
    challan_dict["updated_at"] = challan_dict["updated_at"].isoformat()
    challan_dict["date"] = challan_dict["date"].isoformat()
    if challan_dict.get("payment_date"):
        challan_dict["payment_date"] = challan_dict["payment_date"].isoformat()
    
    await db.challans.insert_one(challan_dict)
    
    # Update driver risk score if driver is linked
    if challan_data.driver_id:
        await update_driver_risk_score(challan_data.driver_id)
    
    return serialize_doc(challan_dict)

@api_router.get("/challans")
async def get_challans(
    vehicle_id: Optional[str] = Query(None, description="Filter by vehicle ID"),
    driver_id: Optional[str] = Query(None, description="Filter by driver ID"),
    status: Optional[str] = Query(None, description="Filter by status (Paid/Unpaid)"),
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(100, ge=1, le=1000, description="Number of records to return"),
    current_user: dict = Depends(get_current_user)
):
    """
    Get challans with optional filters for vehicle_id, driver_id, and status
    """
    query = {"is_deleted": False}
    
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    if driver_id:
        query["driver_id"] = driver_id
    
    if status:
        query["status"] = status
    
    # Get total count for pagination
    total_count = await db.challans.count_documents(query)
    
    # Get paginated results
    challans = await db.challans.find(
        query, 
        {"_id": 0}
    ).sort("date", -1).skip(skip).limit(limit).to_list(limit)
    
    # Calculate summary stats
    unpaid_query = {**query, "status": "Unpaid"}
    unpaid_count = await db.challans.count_documents(unpaid_query)
    
    unpaid_pipeline = [
        {"$match": unpaid_query},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    unpaid_result = await db.challans.aggregate(unpaid_pipeline).to_list(1)
    unpaid_amount = unpaid_result[0]["total"] if unpaid_result else 0
    
    return {
        "data": challans,
        "total": total_count,
        "skip": skip,
        "limit": limit,
        "summary": {
            "unpaid_count": unpaid_count,
            "unpaid_amount": unpaid_amount
        }
    }

@api_router.put("/challans/{challan_id}/pay")
async def pay_challan(
    challan_id: str, 
    payment_date: datetime, 
    current_user: dict = Depends(get_current_user)
):
    result = await db.challans.update_one(
        {"id": challan_id, "is_deleted": False},
        {
            "$set": {
                "status": "Paid",
                "payment_date": payment_date.isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Challan not found")

    # Get the challan to update driver risk score
    challan = await db.challans.find_one({"id": challan_id})
    if challan and challan.get("driver_id"):
        await update_driver_risk_score(challan["driver_id"])

    return {"message": "Challan marked as paid"}

@api_router.delete("/challans/{challan_id}")
async def delete_challan(challan_id: str, current_user: dict = Depends(get_current_user)):
    # Get the challan before deleting to update driver risk score
    challan = await db.challans.find_one({"id": challan_id})
    
    result = await db.challans.update_one(
        {"id": challan_id},
        {
            "$set": {
                "is_deleted": True,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Challan not found")
    
    # Update driver risk score if driver was linked
    if challan and challan.get("driver_id"):
        await update_driver_risk_score(challan["driver_id"])

    return {"message": "Challan deleted"}

@api_router.post("/upload-challan")
async def upload_challan(file: UploadFile = File(...)):

    try:

        if file.content_type.startswith("image"):

            result = cloudinary.uploader.upload(
                file.file,
                folder="challans",
                resource_type="image",
                transformation=[
                    {
                        "quality": "auto",
                        "fetch_format": "auto",
                        "width": 1000,
                        "crop": "limit"
                    }
                ]
            )

        else:

            result = cloudinary.uploader.upload(
                file.file,
                folder="challans",
                resource_type="raw"
            )

        return {
            "url": result["secure_url"],
            "public_id": result["public_id"]
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/export/challans/excel")
async def export_challans_excel(
    vehicle_id: Optional[str] = None,
    driver_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Export challans to a professional Excel file with filtering options
    """
    # Build query
    query = {"is_deleted": False}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    if driver_id:
        query["driver_id"] = driver_id
    if status:
        query["status"] = status
    
    # Get all challans
    challans = await db.challans.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    if not challans:
        raise HTTPException(status_code=404, detail="No challans found to export")
    
    # Get vehicle information for reference
    vehicle_ids = list(set(c["vehicle_id"] for c in challans))
    vehicles = await db.vehicles.find(
        {"id": {"$in": vehicle_ids}},
        {"_id": 0, "id": 1, "registration_number": 1, "brand": 1, "model": 1}
    ).to_list(100)
    
    vehicle_map = {v["id"]: v for v in vehicles}
    
    # Get driver information for reference
    driver_ids = list(set(c.get("driver_id") for c in challans if c.get("driver_id")))
    drivers = await db.drivers.find(
        {"id": {"$in": driver_ids}},
        {"_id": 0, "id": 1, "full_name": 1}
    ).to_list(100)
    
    driver_map = {d["id"]: d for d in drivers}
    
    # Prepare data for Excel
    excel_data = []
    for challan in challans:
        vehicle = vehicle_map.get(challan["vehicle_id"], {})
        driver = driver_map.get(challan.get("driver_id"), {}) if challan.get("driver_id") else None
        
        # Parse dates
        challan_date = datetime.fromisoformat(challan["date"]) if isinstance(challan["date"], str) else challan["date"]
        payment_date = None
        if challan.get("payment_date"):
            payment_date = datetime.fromisoformat(challan["payment_date"]) if isinstance(challan["payment_date"], str) else challan["payment_date"]
        
        excel_data.append({
            "Challan Number": challan.get("challan_number", ""),
            "Registration Number": vehicle.get("registration_number", "Unknown"),
            "Vehicle Model": f"{vehicle.get('brand', '')} {vehicle.get('model', '')}".strip(),
            "Driver Name": driver.get("full_name", "N/A") if driver else "N/A",
            "Violation Type": challan.get("violation_type", ""),
            "Amount (Rs)": challan.get("amount", 0),
            "Date": challan_date.strftime("%Y-%m-%d") if challan_date else "",
            "Location": challan.get("location", ""),
            "Status": challan.get("status", ""),
            "Payment Date": payment_date.strftime("%Y-%m-%d") if payment_date else "",
            "Phone Number": challan.get("phone_number", ""),
            "Has Proof": "Yes" if challan.get("proof_url") else "No",
            "Created At": challan.get("created_at", ""),
        })
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    # Use openpyxl engine
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Challans', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Challans']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        total_amount = sum(c.get("amount", 0) for c in challans)
        paid_amount = sum(c.get("amount", 0) for c in challans if c.get("status") == "Paid")
        unpaid_amount = sum(c.get("amount", 0) for c in challans if c.get("status") == "Unpaid")
        
        summary_data = {
            'Metric': [
                'Total Challans',
                'Paid Challans',
                'Unpaid Challans',
                'Total Amount (Rs)',
                'Paid Amount (Rs)',
                'Unpaid Amount (Rs)',
                'Average Fine (Rs)',
                'Vehicles Involved',
                'Drivers Involved',
                'Export Date'
            ],
            'Value': [
                len(challans),
                len([c for c in challans if c.get("status") == "Paid"]),
                len([c for c in challans if c.get("status") == "Unpaid"]),
                f"₹{total_amount:,.2f}",
                f"₹{paid_amount:,.2f}",
                f"₹{unpaid_amount:,.2f}",
                f"₹{total_amount/len(challans):,.2f}" if challans else "₹0.00",
                len(vehicle_ids),
                len([d for d in driver_ids if d]),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Auto-adjust summary sheet columns
        summary_sheet = writer.sheets['Summary']
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 20
    
    output.seek(0)
    
    # Generate filename based on filters
    filename_parts = ["challans"]
    if vehicle_id:
        vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "registration_number": 1})
        if vehicle:
            filename_parts.append(vehicle["registration_number"].replace(" ", "_"))
    if driver_id:
        driver = await db.drivers.find_one({"id": driver_id}, {"_id": 0, "full_name": 1})
        if driver:
            filename_parts.append(driver["full_name"].replace(" ", "_"))
    if status:
        filename_parts.append(status.lower())
    
    filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
    filename = "_".join(filename_parts) + ".xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/export/challans/current-view/excel")
async def export_challans_current_view_excel(
    filter_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Export the current filtered view of challans to Excel
    """
    challans = filter_data.get("challans", [])
    vehicle_map = filter_data.get("vehicle_map", {})
    driver_map = filter_data.get("driver_map", {})
    
    if not challans:
        raise HTTPException(status_code=404, detail="No challans to export")
    
    # Prepare data for Excel
    excel_data = []
    for challan in challans:
        vehicle = vehicle_map.get(challan.get("vehicle_id"), {})
        driver = driver_map.get(challan.get("driver_id"), {}) if challan.get("driver_id") else None
        
        # Parse dates
        challan_date = datetime.fromisoformat(challan["date"]) if isinstance(challan["date"], str) else challan["date"]
        payment_date = None
        if challan.get("payment_date"):
            payment_date = datetime.fromisoformat(challan["payment_date"]) if isinstance(challan["payment_date"], str) else challan["payment_date"]
        
        excel_data.append({
            "Challan Number": challan.get("challan_number", ""),
            "Registration Number": vehicle.get("registration_number", "Unknown"),
            "Vehicle Model": f"{vehicle.get('brand', '')} {vehicle.get('model', '')}".strip(),
            "Driver Name": driver.get("full_name", "N/A") if driver else "N/A",
            "Violation Type": challan.get("violation_type", ""),
            "Amount (Rs)": challan.get("amount", 0),
            "Date": challan_date.strftime("%Y-%m-%d") if challan_date else "",
            "Location": challan.get("location", ""),
            "Status": challan.get("status", ""),
            "Payment Date": payment_date.strftime("%Y-%m-%d") if payment_date else "",
            "Phone Number": challan.get("phone_number", ""),
            "Has Proof": "Yes" if challan.get("proof_url") else "No",
        })
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Challans', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Challans']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=challans_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )

# FASTag Balance and Transaction Routes

@api_router.post("/vehicles/{vehicle_id}/fastag-balance")
async def get_vehicle_fastag_balance(
    vehicle_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Fetch FASTag balance and details for a vehicle
    """
    try:
        # Find the vehicle
        vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        # Clean the registration number - remove hyphens and spaces
        registration_number = vehicle["registration_number"].upper().replace("-", "").replace(" ", "")
        
        logger.info(f"Fetching FASTag balance for vehicle: {registration_number} (original: {vehicle['registration_number']})")
        
        # Call Surepass API
        result = await surepass_service.fetch_fastag_details(registration_number)
        
        if not result["success"]:
            logger.error(f"FASTag API error: {result.get('error')}")
            raise HTTPException(status_code=400, detail=result.get("error", "Failed to fetch FASTag details"))
        
        fastag_data = result["data"]
        
        # Log the API call
        try:
            log_entry = {
                "id": str(uuid.uuid4()),
                "vehicle_id": vehicle_id,
                "registration_number": registration_number,
                "request_timestamp": datetime.now(timezone.utc).isoformat(),
                "response_data": fastag_data,
                "is_successful": True,
                "created_by": current_user["user_id"]
            }
            await db.fastag_verification_logs.insert_one(log_entry)
        except Exception as e:
            logger.warning(f"Could not log FASTag verification: {e}")
        
        # Update vehicle with latest FASTag information
        update_data = {}
        if fastag_data.get("available_balance") is not None:
            try:
                update_data["fastag_balance"] = float(fastag_data["available_balance"])
            except:
                pass
        if fastag_data.get("bank_name"):
            update_data["fastag_company"] = fastag_data["bank_name"]
        if fastag_data.get("tag_status"):
            update_data["fastag_status"] = fastag_data["tag_status"]
        
        if update_data:
            await db.vehicles.update_one(
                {"id": vehicle_id},
                {"$set": {
                    **update_data,
                    "fastag_last_synced": datetime.now(timezone.utc).isoformat()
                }}
            )
        
        return {
            "success": True,
            "vehicle_id": vehicle_id,
            "registration_number": registration_number,
            "fastag_data": fastag_data,
            "message": "FASTag details fetched successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in get_vehicle_fastag_balance: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@api_router.post("/vehicles/{vehicle_id}/fastag-transactions")
async def get_vehicle_fastag_transactions(
    vehicle_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    Fetch FASTag transaction history for a vehicle
    """
    try:
        # Find the vehicle
        vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False})
        if not vehicle:
            raise HTTPException(status_code=404, detail="Vehicle not found")
        
        # Clean the registration number - remove hyphens and spaces
        registration_number = vehicle["registration_number"].upper().replace("-", "").replace(" ", "")
        
        logger.info(f"Fetching FASTag transactions for vehicle: {registration_number} (original: {vehicle['registration_number']})")
        
        # Call Surepass API
        result = await surepass_service.fetch_fastag_transactions(registration_number)
        
        if not result["success"]:
            logger.error(f"FASTag transaction API error: {result.get('error')}")
            raise HTTPException(status_code=400, detail=result.get("error", "Failed to fetch transactions"))
        
        transactions_data = result["data"]
        
        # Log the API call
        try:
            log_entry = {
                "id": str(uuid.uuid4()),
                "vehicle_id": vehicle_id,
                "registration_number": registration_number,
                "request_timestamp": datetime.now(timezone.utc).isoformat(),
                "response_data": transactions_data,
                "is_successful": True,
                "created_by": current_user["user_id"]
            }
            await db.fastag_verification_logs.insert_one(log_entry)
        except Exception as e:
            logger.warning(f"Could not log FASTag transaction: {e}")
        
        return {
            "success": True,
            "vehicle_id": vehicle_id,
            "registration_number": registration_number,
            "transactions": transactions_data.get("transactions", []),
            "transaction_count": transactions_data.get("transaction_count", 0),
            "tag_id": transactions_data.get("tag_id"),
            "bank_name": transactions_data.get("bank_name"),
            "status": transactions_data.get("status"),
            "message": "FASTag transactions fetched successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in get_vehicle_fastag_transactions: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")
# ==================== SERVICE RECORD ROUTES ====================

@api_router.post("/service-records")
async def create_service_record(service_data: ServiceRecordCreate, current_user: dict = Depends(get_current_user)):
    service_id = str(uuid.uuid4())
    service = ServiceRecord(id=service_id, **service_data.model_dump(), created_by=current_user["user_id"])
    
    service_dict = service.model_dump()
    service_dict["created_at"] = service_dict["created_at"].isoformat()
    service_dict["updated_at"] = service_dict["updated_at"].isoformat()
    service_dict["service_date"] = service_dict["service_date"].isoformat()
    if service_dict.get("next_service_due_date"):
        service_dict["next_service_due_date"] = service_dict["next_service_due_date"].isoformat()
    
    await db.service_records.insert_one(service_dict)
    return serialize_doc(service_dict)

@api_router.get("/service-records")
async def get_service_records(
    vehicle_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": False}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    records = await db.service_records.find(query, {"_id": 0}).sort("service_date", -1).skip(skip).limit(limit).to_list(limit)
    count = await db.service_records.count_documents(query)
    return {"data": records, "total": count}

@api_router.get("/check-phone-usage")
async def check_phone_usage(
    vehicle_id: str,
    phone_number: str,
    type: str,
    current_user: dict = Depends(get_current_user)
):
    
    if type == "document":

        docs = await db.vehicle_documents.find(
            {
                "vehicle_id": vehicle_id,
                "phone_number": phone_number,
                "is_deleted": False
            },
            {"_id": 0}
        ).to_list(50)

        used_in = []

        for d in docs:
            if d["document_type"] == "Custom":
                used_in.append(d.get("custom_document_name"))
            else:
                used_in.append(d["document_type"])

        return {"used_in": used_in}

    elif type == "challan":

        challans = await db.challans.find(
            {
                "vehicle_id": vehicle_id,
                "phone_number": phone_number,
                "is_deleted": False
            },
            {"_id": 0}
        ).to_list(50)

        used_in = [c["violation_type"] for c in challans]

        return {"used_in": used_in}

    return {"used_in": []}

@api_router.post("/check-phone-usage")
async def check_phone_usage(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """
    Check if a phone number is already used in any tax/bill for a specific property
    """
    property_id = data.get("property_id")
    phone_number = data.get("phone_number")
    exclude_id = data.get("exclude_id")
    exclude_type = data.get("exclude_type")
    
    if not property_id or not phone_number:
        return {"used_in": []}
    
    used_in = []
    
    # Check Property Taxes (existing code)
    tax_query = {
        "property_id": property_id,
        "phone_number": phone_number,
        "is_deleted": False
    }
    if exclude_id and exclude_type == "property-tax":
        tax_query["id"] = {"$ne": exclude_id}
    
    property_taxes = await db.property_taxes.find(
        tax_query,
        {"_id": 0, "id": 1, "tax_type": 1, "custom_tax_name": 1}
    ).to_list(100)
    
    for tax in property_taxes:
        name = tax.get("custom_tax_name") if tax.get("tax_type") == "Other" else tax.get("tax_type")
        used_in.append({
            "type": "property-tax",
            "name": name,
            "id": tax["id"]
        })
    
    # Check Electricity Bills (existing code)
    elec_query = {
        "property_id": property_id,
        "phone_number": phone_number,
        "is_deleted": False
    }
    if exclude_id and exclude_type == "electricity":
        elec_query["id"] = {"$ne": exclude_id}
    
    electricity_bills = await db.electricity_bills.find(
        elec_query,
        {"_id": 0, "id": 1, "billing_period_start": 1}
    ).to_list(100)
    
    for bill in electricity_bills:
        period = datetime.fromisoformat(bill["billing_period_start"]).strftime("%b %Y")
        used_in.append({
            "type": "electricity",
            "name": f"Electricity Bill - {period}",
            "id": bill["id"]
        })
    
    # Check Gas Bills (existing code)
    gas_query = {
        "property_id": property_id,
        "phone_number": phone_number,
        "is_deleted": False
    }
    if exclude_id and exclude_type == "gas":
        gas_query["id"] = {"$ne": exclude_id}
    
    gas_bills = await db.gas_bills.find(
        gas_query,
        {"_id": 0, "id": 1, "billing_period_start": 1, "vendor": 1}
    ).to_list(100)
    
    for bill in gas_bills:
        period = datetime.fromisoformat(bill["billing_period_start"]).strftime("%b %Y")
        used_in.append({
            "type": "gas",
            "name": f"Gas Bill ({bill['vendor']}) - {period}",
            "id": bill["id"]
        })
    
    # Check Water Bills (NEW CODE)
    water_query = {
        "property_id": property_id,
        "phone_number": phone_number,
        "is_deleted": False
    }
    if exclude_id and exclude_type == "water":
        water_query["id"] = {"$ne": exclude_id}
    
    water_bills = await db.water_bills.find(
        water_query,
        {"_id": 0, "id": 1, "billing_period_start": 1}
    ).to_list(100)
    
    for bill in water_bills:
        period = datetime.fromisoformat(bill["billing_period_start"]).strftime("%b %Y")
        used_in.append({
            "type": "water",
            "name": f"Water Bill - {period}",
            "id": bill["id"]
        })
    
    return {"used_in": used_in}

# ==================== GPS & TELEMATICS ROUTES ====================

@api_router.post("/gps-devices")
async def create_gps_device(device_data: GPSDeviceCreate, current_user: dict = Depends(get_current_user)):
    device_id = str(uuid.uuid4())
    device = GPSDevice(id=device_id, **device_data.model_dump(), created_by=current_user["user_id"])
    
    device_dict = device.model_dump()
    device_dict["created_at"] = device_dict["created_at"].isoformat()
    device_dict["updated_at"] = device_dict["updated_at"].isoformat()
    
    await db.gps_devices.insert_one(device_dict)
    return serialize_doc(device_dict)

@api_router.get("/gps-devices")
async def get_gps_devices(current_user: dict = Depends(get_current_user)):
    devices = await db.gps_devices.find({"is_deleted": False}, {"_id": 0}).to_list(1000)
    return {"data": devices}

@api_router.post("/gps-devices/{device_id}/simulate")
async def simulate_gps_data(
    device_id: str,
    start_time: datetime,
    end_time: datetime,
    current_user: dict = Depends(get_current_user)
):
    """Generate and store simulated GPS data for a device"""
    device = await db.gps_devices.find_one({"id": device_id, "is_deleted": False})
    if not device:
        raise HTTPException(status_code=404, detail="GPS device not found")
    
    # Generate location data
    locations = await telematics_service.get_location_history(device_id, start_time, end_time)
    
    # Store in database
    for loc in locations:
        loc_id = str(uuid.uuid4())
        log = GPSLocationLog(
            id=loc_id,
            gps_device_id=device_id,
            vehicle_id=device["vehicle_id"],
            **loc
        )
        log_dict = log.model_dump()
        log_dict["timestamp"] = log_dict["timestamp"].isoformat()
        await db.gps_location_logs.insert_one(log_dict)
    
    return {"message": f"Generated {len(locations)} location points", "count": len(locations)}

@api_router.get("/gps-location-logs")
async def get_gps_location_logs(
    vehicle_id: Optional[str] = None,
    start_time: Optional[datetime] = None,
    end_time: Optional[datetime] = None,
    skip: int = 0,
    limit: int = 500,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    if start_time:
        query["timestamp"] = {"$gte": start_time.isoformat()}
    if end_time:
        if "timestamp" not in query:
            query["timestamp"] = {}
        query["timestamp"]["$lte"] = end_time.isoformat()
    
    logs = await db.gps_location_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    count = await db.gps_location_logs.count_documents(query)
    return {"data": logs, "total": count}

@api_router.post("/trips/calculate")
async def calculate_trip(
    vehicle_id: str,
    start_time: datetime,
    end_time: datetime,
    fuel_price_per_liter: float = 100.0,
    current_user: dict = Depends(get_current_user)
):
    """Calculate trip metrics from GPS logs"""
    # Get vehicle
    vehicle = await db.vehicles.find_one({"id": vehicle_id, "is_deleted": False})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    
    # Get GPS device
    device = await db.gps_devices.find_one({"vehicle_id": vehicle_id, "is_deleted": False})
    if not device:
        raise HTTPException(status_code=404, detail="GPS device not found for vehicle")
    
    # Get location logs
    logs = await db.gps_location_logs.find({
        "vehicle_id": vehicle_id,
        "timestamp": {"$gte": start_time.isoformat(), "$lte": end_time.isoformat()}
    }, {"_id": 0}).sort("timestamp", 1).to_list(10000)
    
    if not logs:
        raise HTTPException(status_code=404, detail="No GPS data found for this time range")
    
    # Parse timestamps
    for log in logs:
        if isinstance(log["timestamp"], str):
            log["timestamp"] = datetime.fromisoformat(log["timestamp"])
    
    # Calculate trip metrics
    metrics = telematics_service.calculate_trip_metrics(logs)
    
    # Calculate fuel consumption
    fuel_data = telematics_service.calculate_fuel_consumption(
        metrics["distance_km"],
        vehicle["average_kmpl"],
        fuel_price_per_liter
    )
    
    # Create trip record
    trip_id = str(uuid.uuid4())
    trip = Trip(
        id=trip_id,
        vehicle_id=vehicle_id,
        gps_device_id=device["id"],
        start_time=logs[0]["timestamp"],
        end_time=logs[-1]["timestamp"],
        start_location={"lat": logs[0]["latitude"], "lng": logs[0]["longitude"]},
        end_location={"lat": logs[-1]["latitude"], "lng": logs[-1]["longitude"]},
        distance_km=metrics["distance_km"],
        idle_time_minutes=metrics["idle_time_minutes"],
        average_speed=metrics["average_speed"],
        max_speed=metrics["max_speed"],
        fuel_consumed_liters=fuel_data["fuel_consumed_liters"],
        fuel_cost=fuel_data["fuel_cost"],
        created_by=current_user["user_id"]
    )
    
    trip_dict = trip.model_dump()
    trip_dict["created_at"] = trip_dict["created_at"].isoformat()
    trip_dict["updated_at"] = trip_dict["updated_at"].isoformat()
    trip_dict["start_time"] = trip_dict["start_time"].isoformat()
    trip_dict["end_time"] = trip_dict["end_time"].isoformat()
    
    await db.trips.insert_one(trip_dict)
    return serialize_doc(trip_dict)

@api_router.get("/trips")
async def get_trips(
    vehicle_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": False}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    trips = await db.trips.find(query, {"_id": 0}).sort("start_time", -1).skip(skip).limit(limit).to_list(limit)
    count = await db.trips.count_documents(query)
    return {"data": trips, "total": count}

# ==================== DRIVER ROUTES ====================

@api_router.post("/drivers")
async def create_driver(driver_data: DriverCreate, current_user: dict = Depends(get_current_user)):
    driver_id = str(uuid.uuid4())
    driver = Driver(id=driver_id, **driver_data.model_dump(), created_by=current_user["user_id"])
    
    driver_dict = driver.model_dump()
    driver_dict["created_at"] = driver_dict["created_at"].isoformat()
    driver_dict["updated_at"] = driver_dict["updated_at"].isoformat()
    driver_dict["license_expiry"] = driver_dict["license_expiry"].isoformat()
    
    await db.drivers.insert_one(driver_dict)
    return serialize_doc(driver_dict)

@api_router.get("/drivers")
async def get_drivers(
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    drivers = await db.drivers.find({"is_deleted": False}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    count = await db.drivers.count_documents({"is_deleted": False})
    return {"data": drivers, "total": count}

@api_router.put("/drivers/{driver_id}")
async def update_driver(driver_id: str, driver_data: DriverCreate, current_user: dict = Depends(get_current_user)):
    update_data = driver_data.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["license_expiry"] = update_data["license_expiry"].isoformat()
    
    result = await db.drivers.update_one({"id": driver_id, "is_deleted": False}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Driver not found")
    return {"message": "Driver updated"}

async def update_driver_risk_score(driver_id: str):
    """Calculate and update driver risk score based on violations and accidents"""
    if not driver_id:
        return
    
    # Get all unpaid challans for this driver
    unpaid_challans = await db.challans.count_documents({
        "driver_id": driver_id, 
        "status": "Unpaid",
        "is_deleted": False
    })
    
    # Get all paid challans for this driver
    paid_challans = await db.challans.count_documents({
        "driver_id": driver_id, 
        "status": "Paid",
        "is_deleted": False
    })
    
    # Get accident count
    accident_count = await db.accidents.count_documents({
        "driver_id": driver_id, 
        "is_deleted": False
    })
    
    # Calculate total challan amount
    challan_pipeline = [
        {"$match": {"driver_id": driver_id, "is_deleted": False}},
        {"$group": {"_id": None, "total_amount": {"$sum": "$amount"}}}
    ]
    challan_result = await db.challans.aggregate(challan_pipeline).to_list(1)
    total_challan_amount = challan_result[0]["total_amount"] if challan_result else 0
    
    # Calculate risk score based on multiple factors
    # Base score from number of challans
    base_score = (unpaid_challans * 15) + (paid_challans * 5)
    
    # Additional weight for high-value challans
    amount_score = min(total_challan_amount / 1000, 50)  # Max 50 points from amount
    
    # Accident score
    accident_score = accident_count * 50
    
    risk_score = base_score + amount_score + accident_score
    
    await db.drivers.update_one(
        {"id": driver_id}, 
        {"$set": {
            "risk_score": risk_score,
            "total_challans": unpaid_challans + paid_challans,
            "unpaid_challans": unpaid_challans,
            "total_challan_amount": total_challan_amount
        }}
    )

# Add these endpoints to your server.py

@api_router.get("/warnings/vehicle/{vehicle_id}")
async def get_vehicle_warnings(
    vehicle_id: str,
    date: Optional[str] = None,
    violation_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Get warnings for a vehicle including:
    - Multiple challans on same day
    - Pending challans count and amount
    - Similar violation patterns
    """
    query = {"vehicle_id": vehicle_id, "is_deleted": False}
    
    # Get all challans for this vehicle
    vehicle_challans = await db.challans.find(
        query, 
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    warnings = []
    
    # Check for multiple challans on same day
    if date:
        target_date = datetime.fromisoformat(date).date()
        same_day_challans = [
            c for c in vehicle_challans 
            if datetime.fromisoformat(c["date"]).date() == target_date
        ]
        
        if len(same_day_challans) > 0:
            warnings.append({
                "type": "multiple_same_day",
                "severity": "warning",
                "message": f"Vehicle has {len(same_day_challans)} challan(s) on this date",
                "details": [
                    {
                        "time": c["date"],
                        "type": c["violation_type"],
                        "amount": c["amount"]
                    } for c in same_day_challans
                ]
            })
    
    # Check pending challans
    pending_challans = [c for c in vehicle_challans if c["status"] == "Unpaid"]
    if len(pending_challans) >= 2:
        total_amount = sum(c["amount"] for c in pending_challans)
        warnings.append({
            "type": "pending",
            "severity": "warning",
            "message": f"Vehicle has {len(pending_challans)} pending challans",
            "total_amount": total_amount
        })
    
    # Check similar violations
    if violation_type:
        similar = [
            c for c in vehicle_challans 
            if c["violation_type"].lower() == violation_type.lower()
            and c["status"] == "Unpaid"
        ]
        if similar:
            warnings.append({
                "type": "repeat",
                "severity": "info",
                "message": f"Similar violation recorded on {similar[0]['date'][:10]}"
            })
    
    return {"warnings": warnings}


@api_router.get("/warnings/driver/{driver_id}")
async def get_driver_warnings(
    driver_id: str,
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """
    Get warnings for a driver including:
    - Multiple challans on same day (with vehicle details)
    - Multiple challans in last 7 days
    - Total challans and unpaid amount
    """
    query = {"driver_id": driver_id, "is_deleted": False}
    
    # Get all challans for this driver
    driver_challans = await db.challans.find(
        query, 
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Get vehicle details for reference
    vehicle_ids = list(set(c["vehicle_id"] for c in driver_challans))
    vehicles = await db.vehicles.find(
        {"id": {"$in": vehicle_ids}, "is_deleted": False},
        {"_id": 0, "id": 1, "registration_number": 1}
    ).to_list(100)
    
    vehicle_map = {v["id"]: v["registration_number"] for v in vehicles}
    
    warnings = []
    
    # Check for multiple challans on same day
    if date:
        target_date = datetime.fromisoformat(date).date()
        same_day_challans = [
            c for c in driver_challans 
            if datetime.fromisoformat(c["date"]).date() == target_date
        ]
        
        if len(same_day_challans) > 0:
            warnings.append({
                "type": "driver_multiple_same_day",
                "severity": "warning",
                "message": f"Driver has {len(same_day_challans)} challan(s) on this date",
                "details": [
                    {
                        "vehicle": vehicle_map.get(c["vehicle_id"], "Unknown"),
                        "time": c["date"],
                        "type": c["violation_type"],
                        "amount": c["amount"]
                    } for c in same_day_challans
                ]
            })
    
    # Check last 7 days activity
    seven_days_ago = datetime.now(timezone.utc) - timedelta(days=7)
    week_challans = [
        c for c in driver_challans 
        if datetime.fromisoformat(c["date"]) >= seven_days_ago
    ]
    
    if len(week_challans) >= 2:
        vehicle_summary = {}
        for c in week_challans:
            reg = vehicle_map.get(c["vehicle_id"], "Unknown")
            vehicle_summary[reg] = vehicle_summary.get(reg, 0) + 1
        
        warnings.append({
            "type": "driver_week_multiple",
            "severity": "warning",
            "message": f"Driver has {len(week_challans)} challans in last 7 days",
            "details": {
                "vehicles": ", ".join([f"{v} ({c})" for v, c in vehicle_summary.items()]),
                "count": len(week_challans)
            }
        })
    
    # Check total statistics
    total_challans = len(driver_challans)
    if total_challans >= 3:
        warnings.append({
            "type": "frequent",
            "severity": "warning",
            "message": f"Driver has {total_challans} total challans"
        })
    
    unpaid_amount = sum(c["amount"] for c in driver_challans if c["status"] == "Unpaid")
    if unpaid_amount > 5000:
        warnings.append({
            "type": "amount",
            "severity": "warning",
            "message": f"Driver has Rs {unpaid_amount:,.2f} in unpaid challans"
        })
    
    return {"warnings": warnings}    
# ==================== ACCIDENT & CLAIM ROUTES ====================

@api_router.post("/accidents")
async def create_accident(accident_data: AccidentCreate, current_user: dict = Depends(get_current_user)):
    accident_id = str(uuid.uuid4())
    accident = Accident(id=accident_id, **accident_data.model_dump(), created_by=current_user["user_id"])
    
    accident_dict = accident.model_dump()
    accident_dict["created_at"] = accident_dict["created_at"].isoformat()
    accident_dict["updated_at"] = accident_dict["updated_at"].isoformat()
    accident_dict["accident_date"] = accident_dict["accident_date"].isoformat()
    
    await db.accidents.insert_one(accident_dict)
    
    # Update driver risk score if driver is linked
    if accident_data.driver_id:
        await update_driver_risk_score(accident_data.driver_id)
    
    return serialize_doc(accident_dict)

@api_router.get("/accidents")
async def get_accidents(
    vehicle_id: Optional[str] = None,
    driver_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": False}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    if driver_id:
        query["driver_id"] = driver_id
    
    accidents = await db.accidents.find(query, {"_id": 0}).sort("accident_date", -1).skip(skip).limit(limit).to_list(limit)
    count = await db.accidents.count_documents(query)
    return {"data": accidents, "total": count}

@api_router.put("/accidents/{accident_id}")
async def update_accident(accident_id: str, accident_data: AccidentCreate, current_user: dict = Depends(get_current_user)):
    update_data = accident_data.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    update_data["accident_date"] = update_data["accident_date"].isoformat()
    
    result = await db.accidents.update_one({"id": accident_id, "is_deleted": False}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Accident not found")
    return {"message": "Accident updated"}

# ==================== NOTIFICATION ROUTES ====================

@api_router.get("/notifications")
async def get_notifications(
    unread_only: bool = False,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user)
):
    query = {"is_deleted": False}
    if unread_only:
        query["is_read"] = False
    
    notifications = await db.notifications.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    count = await db.notifications.count_documents(query)
    unread_count = await db.notifications.count_documents({"is_read": False, "is_deleted": False})
    
    return {"data": notifications, "total": count, "unread_count": unread_count}

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {"is_read": True, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    return {"message": "Notification marked as read"}

# ==================== ANALYTICS & DASHBOARD ROUTES ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    """Get enterprise dashboard statistics"""
    now = datetime.now(timezone.utc)
    
    # Property stats
    total_properties = await db.properties.count_documents({"is_deleted": False})
    
    # Vehicle stats
    total_vehicles = await db.vehicles.count_documents({"is_deleted": False})
    
    # Unpaid bills
    unpaid_electricity = await db.electricity_bills.count_documents({"status": "Unpaid", "is_deleted": False})
    unpaid_gas = await db.gas_bills.count_documents({"status": "Unpaid", "is_deleted": False})
    unpaid_water = await db.water_bills.count_documents({"status": "Unpaid", "is_deleted": False})
    total_unpaid_bills = unpaid_electricity + unpaid_gas + unpaid_water
    
    # Expired taxes
    expired_taxes = await db.property_taxes.count_documents({
        "expiry_date": {"$lt": now.isoformat()},
        "status": "Unpaid",
        "is_deleted": False
    })
    
    # Expiring documents (next 30 days)
    expiring_threshold = now + timedelta(days=30)
    expiring_docs = await db.vehicle_documents.count_documents({
        "expiry_date": {"$lte": expiring_threshold.isoformat(), "$gte": now.isoformat()},
        "is_current": True,
        "is_deleted": False
    })
    
    # Unpaid challans
    unpaid_challans_pipeline = [
        {"$match": {"status": "Unpaid", "is_deleted": False}},
        {"$group": {"_id": None, "count": {"$sum": 1}, "total": {"$sum": "$amount"}}}
    ]
    unpaid_challans_result = await db.challans.aggregate(unpaid_challans_pipeline).to_list(1)
    unpaid_challans_count = unpaid_challans_result[0]["count"] if unpaid_challans_result else 0
    unpaid_challans_amount = unpaid_challans_result[0]["total"] if unpaid_challans_result else 0
    
    # Total energy consumption (last 30 days)
    thirty_days_ago = now - timedelta(days=30)
    energy_pipeline = [
        {"$match": {
            "billing_period_start": {"$gte": thirty_days_ago.isoformat()},
            "is_deleted": False
        }},
        {"$group": {"_id": None, "total_units": {"$sum": "$units_consumed"}}}
    ]
    energy_result = await db.electricity_bills.aggregate(energy_pipeline).to_list(1)
    total_energy_units = energy_result[0]["total_units"] if energy_result else 0
    
    # Solar generation (last 30 days)
    solar_pipeline = [
        {"$match": {
            "billing_period_start": {"$gte": thirty_days_ago.isoformat()},
            "is_deleted": False
        }},
        {"$group": {"_id": None, "total_generated": {"$sum": "$units_generated"}}}
    ]
    solar_result = await db.solar_meters.aggregate(solar_pipeline).to_list(1)
    total_solar_generated = solar_result[0]["total_generated"] if solar_result else 0
    
    # Total fuel cost (last 30 days)
    fuel_pipeline = [
        {"$match": {
            "start_time": {"$gte": thirty_days_ago.isoformat()},
            "is_deleted": False
        }},
        {"$group": {"_id": None, "total_cost": {"$sum": "$fuel_cost"}, "total_distance": {"$sum": "$distance_km"}}}
    ]
    fuel_result = await db.trips.aggregate(fuel_pipeline).to_list(1)
    total_fuel_cost = fuel_result[0]["total_cost"] if fuel_result else 0
    total_distance = fuel_result[0]["total_distance"] if fuel_result else 0
    
    # Unread notifications
    unread_notifications = await db.notifications.count_documents({"is_read": False, "is_deleted": False})
    
    # Critical alerts
    critical_alerts = await db.notifications.count_documents({
        "severity": "critical",
        "is_read": False,
        "is_deleted": False
    })
    
    # Sustainability metrics
    renewable_percentage = (total_solar_generated / total_energy_units * 100) if total_energy_units > 0 else 0
    co2_saved = total_solar_generated * 0.92  # Approx 0.92 kg CO2 per kWh
    
    return {
        "properties": {
            "total": total_properties
        },
        "vehicles": {
            "total": total_vehicles
        },
        "bills": {
            "unpaid_count": total_unpaid_bills
        },
        "taxes": {
            "expired_count": expired_taxes
        },
        "documents": {
            "expiring_count": expiring_docs
        },
        "challans": {
            "unpaid_count": unpaid_challans_count,
            "unpaid_amount": round(unpaid_challans_amount, 2)
        },
        "energy": {
            "total_consumption_kwh": round(total_energy_units, 2),
            "solar_generation_kwh": round(total_solar_generated, 2),
            "renewable_percentage": round(renewable_percentage, 2)
        },
        "fleet": {
            "total_fuel_cost": round(total_fuel_cost, 2),
            "total_distance_km": round(total_distance, 2),
            "cost_per_km": round(total_fuel_cost / total_distance, 2) if total_distance > 0 else 0
        },
        "notifications": {
            "unread_count": unread_notifications,
            "critical_count": critical_alerts
        },
        "sustainability": {
            "co2_saved_kg": round(co2_saved, 2),
            "renewable_percentage": round(renewable_percentage, 2)
        }
    }

@api_router.get("/analytics/energy-trends")
async def get_energy_trends(
    property_id: Optional[str] = None,
    months: int = 12,
    current_user: dict = Depends(get_current_user)
):
    """Get monthly energy consumption trends"""
    query = {"is_deleted": False}
    if property_id:
        query["property_id"] = property_id
    
    # Get last N months of data
    bills = await db.electricity_bills.find(query, {"_id": 0}).sort("billing_period_start", -1).limit(months).to_list(months)
    
    trends = []
    for bill in reversed(bills):
        if isinstance(bill["billing_period_start"], str):
            period_start = datetime.fromisoformat(bill["billing_period_start"])
        else:
            period_start = bill["billing_period_start"]
        
        trends.append({
            "month": period_start.strftime("%b %Y"),
            "units_consumed": bill["units_consumed"],
            "total_amount": bill["total_amount"]
        })
    
    return {"data": trends}

@api_router.get("/analytics/fuel-efficiency")
async def get_fuel_efficiency(
    vehicle_id: Optional[str] = None,
    months: int = 6,
    current_user: dict = Depends(get_current_user)
):
    """Get fuel efficiency analytics"""
    query = {"is_deleted": False}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    
    # Get trips from last N months
    cutoff_date = datetime.now(timezone.utc) - timedelta(days=months * 30)
    query["start_time"] = {"$gte": cutoff_date.isoformat()}
    
    trips = await db.trips.find(query, {"_id": 0}).sort("start_time", -1).to_list(1000)
    
    # Group by vehicle and month
    vehicle_stats = {}
    for trip in trips:
        vid = trip["vehicle_id"]
        if vid not in vehicle_stats:
            vehicle_stats[vid] = {
                "vehicle_id": vid,
                "total_distance": 0,
                "total_fuel_consumed": 0,
                "total_fuel_cost": 0,
                "trip_count": 0
            }
        
        vehicle_stats[vid]["total_distance"] += trip["distance_km"]
        vehicle_stats[vid]["total_fuel_consumed"] += trip["fuel_consumed_liters"]
        vehicle_stats[vid]["total_fuel_cost"] += trip["fuel_cost"]
        vehicle_stats[vid]["trip_count"] += 1
    
    # Calculate efficiency
    for vid, stats in vehicle_stats.items():
        if stats["total_fuel_consumed"] > 0:
            stats["average_kmpl"] = round(stats["total_distance"] / stats["total_fuel_consumed"], 2)
        else:
            stats["average_kmpl"] = 0
        
        if stats["total_distance"] > 0:
            stats["cost_per_km"] = round(stats["total_fuel_cost"] / stats["total_distance"], 2)
        else:
            stats["cost_per_km"] = 0
    
    return {"data": list(vehicle_stats.values())}

# ==================== KEEP-ALIVE ENDPOINT FOR UPTIMEROBOT ====================

@api_router.get("/ping")
async def ping():
    """
    Simple ping endpoint for UptimeRobot to keep the server alive.
    This endpoint:
    - Responds quickly (no database queries)
    - Returns minimal data
    - Wakes up the server when called
    """
    return {
        "status": "alive",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "service": "Enterprise ERP Backend"
    }

# Optional but recommended - health check with database verification
@api_router.get("/health")
async def health_check():
    """
    Health check endpoint that verifies database connection.
    Use this for more thorough monitoring.
    """
    try:
        # Quick database check
        await db.command("ping")
        db_status = "connected"
    except Exception as e:
        db_status = f"disconnected: {str(e)}"
    
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "database": db_status,
        "service": "Enterprise ERP Backend"
    }    

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logger.info("Enterprise ERP Backend Started")
